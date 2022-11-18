import ast
import os
from math import sqrt
from typing import List, Optional

import numpy as np
from PySide2.QtCore import Qt
from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, borders
import openpyxl
from PySide2.QtWidgets import QWidget, QPushButton, QFileDialog, QDialog, QLabel, QTableWidget, \
    QTableWidgetItem, QMessageBox, QPlainTextEdit, QLineEdit, QComboBox
import pandas as pd
from openpyxl.styles import PatternFill, Font
import seaborn as sns
from openpyxl.utils import get_column_letter

from PreShin.loggers import logger

# 3d landmark 번호, x, y, z 의 모음이 모음이 들어있는 predict, label txt 파일의 폴더들
# grouping 되어 있는 json 파일이 필요함
# 3d 용 json 파일 존재함

def btn_manual_clicked():
    os.startfile(f'{os.getcwd()}/AI_manual.pdf')  # 메뉴얼 오픈


# 전체 df의 전체 평균
def average(df):
    data_sum = df.sum()  # 각각 id의 sum
    id_sum = data_sum.sum()  # data sum 의 sum
    data_count = df.count()  # df_sheet 의 각각 id의 value 개수
    data_count = data_count.sum()  # id의 value 개수 합
    avr = id_sum / data_count  # 전체 평균
    return avr


def messagebox(text: str, i: str):
    signBox = QMessageBox()
    signBox.setWindowTitle(text)
    signBox.setText(i)

    signBox.setIcon(QMessageBox.Information)
    signBox.setStandardButtons(QMessageBox.Ok)
    signBox.exec_()


class PreShin_UI(QWidget):

    def __init__(self):
        super().__init__()

        self.std_outlier = float()
        self.std = float()
        self.lbl_id = Optional[str]
        self.lbl_list = List[str]
        self.pre_list = List[str]
        self.pre_id = Optional[str]
        self.landmark_name_value = list
        self.landmark_value = List[str]
        self.landmark_key = List[str]
        self.id_list = List[str]
        self.loc_xlsx = str()
        self.new_xlsx_outlier = str
        self.new_xlsx = str()
        self.landmark_name = List[str]
        self.number = list
        self.df_result = pd.DataFrame
        self.df_result_outlier = pd.DataFrame
        self.df_sheet2_name = pd.DataFrame
        self.group_num = list

        self.thin_border = Border(left=borders.Side(style='thin'),
                                  right=borders.Side(style='thin'),
                                  top=borders.Side(style='thin'),
                                  bottom=borders.Side(style='thin'))
        self.blue_color = PatternFill(start_color='b3d9ff', end_color='b3d9ff', fill_type='solid')
        self.green_color = PatternFill(start_color='c1f0c1', end_color='c1f0c1', fill_type='solid')
        self.red_color = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
        self.yellow_color = PatternFill(start_color='ffffb3', end_color='ffffb3', fill_type='solid')
        self.yellow_color2 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.gray_color2 = PatternFill(start_color='e0e0eb', end_color='e0e0eb', fill_type='solid')
        self.gray_color = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
        self.blue_color2 = PatternFill(start_color='ccf5ff', end_color='ccf5ff', fill_type='solid')
        self.orange_color = PatternFill(start_color='ff9900', end_color='ff9900', fill_type='solid')
        self.outlier_color = PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid')
        self.white_color = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        self.dialog = QDialog()
        self.initUI()

    def initUI(self):
        logger.info('PreShin_UI open')
        batch = '4'
        rate = '2e-4'
        optimizer = 'adam'
        aug = '0'
        comment = 'write comment'
        safe_zone = '3'
        outlier = '5'

        self.table = QTableWidget(4, 2, self.dialog)
        self.table.setSortingEnabled(False)  # 정렬 기능
        self.table.resizeRowsToContents()
        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(0, 80)  # checkbox 열 폭 강제 조절.
        self.table.setColumnWidth(1, 80)

        self.table.setItem(0, 0, QTableWidgetItem('Batch size'))
        self.table.setItem(0, 1, QTableWidgetItem(batch))
        self.table.setItem(1, 0, QTableWidgetItem('Learning rate'))
        self.table.setItem(1, 1, QTableWidgetItem(rate))
        self.table.setItem(2, 0, QTableWidgetItem('Optimizer'))
        self.table.setItem(2, 1, QTableWidgetItem(optimizer))
        self.table.setItem(3, 0, QTableWidgetItem('Aug'))
        self.table.setItem(3, 1, QTableWidgetItem(aug))

        self.table.setHorizontalHeaderLabels(["Name", "Value"])
        self.table.setGeometry(20, 195, 180, 145)

        btn_pre_path = QPushButton(self.dialog)
        btn_lbl_path = QPushButton(self.dialog)
        btn_export = QPushButton(self.dialog)
        btn_manual = QPushButton(self.dialog)

        btn_pre_path.setText('Predict Path')
        btn_lbl_path.setText('Label Path')
        btn_export.setText('Export Excel')
        btn_manual.setText('Open Manual')

        self.lbl_pre = QLabel(self.dialog)
        self.lbl_lbl = QLabel(self.dialog)
        lbl_unit = QLabel(self.dialog)
        lbl_error = QLabel(self.dialog)
        lbl_outlier = QLabel(self.dialog)

        lbl_comment = QLabel(self.dialog)
        lbl_xlsx_name = QLabel(self.dialog)
        lbl_xlsx = QLabel(self.dialog)
        self.edt_error = QLineEdit(self.dialog)
        self.edt_error.setAlignment(Qt.AlignRight)
        self.edt_outlier = QLineEdit(self.dialog)
        self.edt_outlier.setAlignment(Qt.AlignRight)
        self.edt_xlsx_name = QLineEdit(self.dialog)
        self.edt_xlsx_name.setAlignment(Qt.AlignRight)

        self.combobox()

        self.lbl_outlier_unit.setGeometry(270, 305, 100, 20)
        self.lbl_lbl.setGeometry(130, 35, 230, 20)
        self.lbl_pre.setGeometry(130, 60, 230, 20)
        lbl_unit.setGeometry(220, 200, 100, 20)
        lbl_error.setGeometry(220, 230, 100, 20)
        lbl_outlier.setGeometry(220, 280, 100, 20)
        self.lbl_mm.setGeometry(270, 255, 50, 20)
        lbl_comment.move(20, 90)
        lbl_xlsx_name.move(20, 355)
        lbl_xlsx.move(173, 355)
        self.edt_error.setGeometry(220, 255, 50, 20)
        self.edt_outlier.setGeometry(220, 305, 50, 20)
        self.edt_xlsx_name.setGeometry(70, 350, 103, 20)

        lbl_unit.setText('Unit Setting')
        lbl_error.setText('Error Safe Zone')
        lbl_outlier.setText('outlier')
        lbl_comment.setText('Comment')
        lbl_xlsx_name.setText('파일명 : ')
        lbl_xlsx.setText('.xlsx')

        self.edt_error.setText(safe_zone)
        self.edt_outlier.setText(outlier)
        btn_manual.setGeometry(20, 10, 100, 20)
        btn_lbl_path.setGeometry(20, 35, 100, 20)
        btn_pre_path.setGeometry(20, 60, 100, 20)
        btn_export.setGeometry(220, 345, 120, 30)

        btn_lbl_path.clicked.connect(self.btn_lbl_clicked)
        btn_pre_path.clicked.connect(self.btn_pre_clicked)
        btn_export.clicked.connect(self.btn_export_clicked)
        btn_manual.clicked.connect(btn_manual_clicked)

        self.edt = QPlainTextEdit(self.dialog)
        self.edt.setPlainText(comment)
        self.edt.setGeometry(20, 105, 300, 80)

        self.dialog.setWindowTitle('AI')
        self.dialog.setGeometry(500, 300, 370, 420)
        self.dialog.exec()
        logger.info('PreShin_UI close')

    def combobox(self):
        cb = QComboBox(self.dialog)
        cb.addItem('mm')
        cb.addItem('Pixel')
        self.lbl_outlier_unit = QLabel('mm', self.dialog)
        self.lbl_mm = QLabel('mm', self.dialog)
        cb.setGeometry(290, 200, 70, 20)
        cb.currentTextChanged.connect(self.cb_unit_change)

    # 콤보 박스 변환
    def cb_unit_change(self, text: str):
        self.lbl_outlier_unit.setText(text)
        self.lbl_mm.setText(text)

    def btn_lbl_clicked(self):
        logger.info('label_btn in')
        # landmark.dat, json 먼저 읽고 변환
        # export 에서 하면 안되서 미리 넣음
        self.landmark()

        self.lbl_id = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(),
                                                       QFileDialog.ShowDirsOnly)  # 창 title, 주소 나중에 변경

        # 폴더 경로 입력
        if self.lbl_id != '':
            self.lbl_list = os.listdir(str(self.lbl_id))  # 폴더 경로에 있는 파일 읽기

            if self.lbl_list is not None:  # 빈 폴더가 아닐 때

                for i in range(len(self.lbl_list)):
                    path, ext = os.path.splitext(self.lbl_list[i])  # 경로, 확장자 분리

                    if ext != '.txt' or ext == '':
                        messagebox('Warning', '폴더안 파일의 형식이 올바르지 않습니다. 폴더를 확인하세요.')
                        logger.error('label file format Error')
                        self.lbl_lbl.setText('')
                        break

                    self.lbl_lbl.setText(str(self.lbl_id))
                    # id 안에 있는 landmark 를 landmark.dat 에 있는 num 를 비교후 저장
                    # export 에서 하면 안되서 미리 넣음
                    self.landmark_name = self.compare_landmark(self.lbl_id, self.lbl_list)

            else:
                messagebox('Warning', '폴더안 파일의 형식이 올바르지 않습니다. 폴더를 확인하세요.')
                logger.error('label file format Error')

        else:
            self.lbl_lbl.setText('')
        logger.info('label_btn out')

    def btn_pre_clicked(self):
        logger.info('predict_btn in')
        self.pre_id = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(),
                                                       QFileDialog.ShowDirsOnly)  # 주소 나중에 변경
        if self.pre_id != '':
            self.pre_list = os.listdir(str(self.pre_id))  # 경로에 있는 파일 읽기

            if self.pre_list is not None:  # 빈 폴더가 아닐때

                for i in range(len(self.pre_list)):
                    path, ext = os.path.splitext(self.pre_list[i])  # 경로, 확장자 분리

                    if ext != '.txt' or ext == '':
                        messagebox('Warning', '폴더안 파일의 형식이 올바르지 않습니다. 폴더를 확인하세요.')
                        logger.error('label file format Error')
                        self.lbl_pre.setText('')
                        break

                    self.lbl_pre.setText(str(self.pre_id))
            else:
                messagebox('Warning', '폴더안 파일의 형식이 올바르지 않습니다. 폴더를 확인하세요.')
                logger.error('label file format Error')

        else:
            self.lbl_pre.setText('')  # 껏을때 빈칸
        logger.info('predict_btn out')

    # landmark.dat 구조 변경 후 number - key, name - value 로 지정
    # data 오류 확인
    def landmark(self):

        txt = open(f'{os.getcwd()}/landmark.dat', 'r')

        landmark = txt.readlines()
        landmark_chunk = []

        for line in landmark:
            # split 을 할수 있도록 landmark.dat 구조 파악한 후 변경해서 분리
            # 한줄에 총 12개
            # 1	1	N	V notch of frontal	3	1	0	0	1	0	0	0
            # 총 landmark list 안에 12개의 list 생성
            landmark = line.replace(',', ' ')
            landmark = landmark.replace('\t', ',')
            landmark = landmark.replace('\n', '')
            landmark = landmark.replace('   ', ',')
            landmark = landmark.split(',')

            # 필요한 위치는 2,3번째
            if len(landmark) < 3:
                logger.error('landmark.data format error')
                logger.error(landmark)

            else:
                landmark_chunk.append(landmark)

        txt.close()
        # id : key , number : value 형태 dict 로 만듬
        landmark_dict = {}
        for i in range(len(landmark_chunk) - 1):
            landmark_dict[landmark_chunk[i][2]] = landmark_chunk[i][1]

        # key, value 분리
        self.landmark_key = list(landmark_dict.keys())
        self.landmark_value = list(landmark_dict.values())

        # 2[Sella] 형태 만듬
        self.landmark_name_value = []
        for i in range(len(self.landmark_key)):
            self.landmark_name_value.append(str(self.landmark_value[i]) + '[' + str(self.landmark_key[i]) + ']')

    def open_json(self):
        with open(f'{os.getcwd()}/group_points_preShin.json', 'r') as inf:  # group : { landmark 번호, ...}
            group = ast.literal_eval(inf.read())  # 그룹 포인트 프리신을 dict 로 변환
        return group

    # id 정렬
    def set_pre_lbl(self):
        set_lbl = set(self.lbl_list)
        set_pre = set(self.pre_list)

        id_list = list(set_lbl & set_pre)  # id 두개다 있는 것만 추려냄
        id_list = [b.split('.')[0] for b in id_list]
        id_list.sort()  # id 정렬 set 함수는 정렬 안되서 나옴
        id_list.reverse()
        self.id_list = [(j + '.txt') for j in id_list]

    # label, predict 폴더 비교 없는 파일 출력
    def error_id(self):
        set_lbl = set(self.lbl_list)
        set_pre = set(self.pre_list)
        only_lbl = list(set_lbl - set_pre)  # label 만 있는 파일
        only_pre = list(set_pre - set_lbl)  # predict 만 있는 파일

        if only_lbl == [] and only_lbl == []:
            pass

        else:
            messagebox('Warning', "label 또는 predict 에 존재하지 않는 id가 있습니다.")
            logger.error('label, predict files not matching')
            logger.error(f'label 폴더에 {only_pre} : 파일이 존재하지 않습니다.')
            logger.error(f'Predict 폴더에 {only_lbl} : 파일이 존재하지 않습니다.')

    # id 안에 있는 landmark 를 landmark.dat 에 있는 num 를 비교후 저장
    def compare_landmark(self, lbl_id: str, lbl_list: list):
        # label 폴더의 제일 처음 환자 id를 읽음
        # landmark, x, y, z 형태
        lines_chunk = self.landmark_id_format_change(lbl_id, lbl_list[0])

        # landmark 번호만 따로 저장
        lines_chunk_num = []
        for i in range(len(lines_chunk)):
            lines_chunk_num.append(lines_chunk[i][0])

        set_lines_chunk_num = set(lines_chunk_num)
        set_landmark_value = set(self.landmark_value)
        empty = set_lines_chunk_num - set_landmark_value
        empty_list = list(empty)  # 집합을 만들어 차집합 으로 landmark.dat 에 없는 num 를 찾음

        landmark_name = []  # 빈 리스트 생성
        # landmark 저장
        for i in range(len(lines_chunk)):

            for j in range(len(self.landmark_value)):

                if lines_chunk[i][0] == self.landmark_value[j]:  # 비교후 같은 값을 landmark_name 에 리스트로 추가
                    landmark_name.append(self.landmark_key[j])  # landmark key : id, value : number
                    continue

                elif j > len(self.landmark_value) - 2:

                    for k in range(len(empty_list)):

                        if empty_list[k] == lines_chunk[i][0]:  # 없는 num 와 비교후 같으면 empty 저장
                            landmark_name.append('None')

        return landmark_name

    def id_dataframe(self, lines_chunk: list) -> pd.DataFrame:
        df = pd.DataFrame(lines_chunk, columns=['Landmark_num', 'x', 'y', 'z'])  # label 데이터 프레임
        df['x'] = df['x'].astype(float)  # 타입 변경 안하면 연산 안됨
        df['y'] = df['y'].astype(float)
        df['z'] = df['z'].astype(float)
        df['Landmark_num'] = df['Landmark_num'].astype(int)
        df = df[df >= 0]

        df = df.sort_values(by='Landmark_num')  # 데이터 정렬
        return df

    # [id, x, y, z] 형태 list 로 만듬
    def landmark_id_format_change(self, loc: str, id_list: str) -> List[List[str]]:
        label = open(str(loc + '/' + id_list), "r", encoding="UTF-8")
        id_format = label.readlines()
        lines = []
        for line in id_format:
            line = line.replace("\n", "")
            line = line.split(",")
            if len(line) != 4:
                logger.error('id landmark format error : [id, x, y, z]')
                logger.error(line)
            else:
                lines.append(line)

        lines.sort(key=lambda x: x[0])
        label.close()
        return lines

    def drop_landmark(self, df: pd.DataFrame):
        df.drop('x', axis=1, inplace=True)
        df.drop('y', axis=1, inplace=True)
        df.drop('z', axis=1, inplace=True)  # x,y,z제거

    def btn_export_clicked(self):
        self.landmark()
        logger.info('btn_export_clicked')
        # lbl, pre 둘다 선택
        if self.lbl_lbl.text() != '' and self.lbl_pre.text() != '':

            if self.edt_xlsx_name.text() != '':  # 파일명 입력 했을때

                loc_xlsx = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(),
                                                            QFileDialog.ShowDirsOnly)
                if loc_xlsx != '':  # 폴더 선택 했을때
                    file = os.listdir(loc_xlsx)  # 엑셀 저장 위치에 있는 파일 읽기

                    if fr'{self.edt_xlsx_name.text()}_folder' not in file:  # 동일한 파일명이 없을때

                        # 저장할 폴더 생성
                        self.loc_xlsx = loc_xlsx + fr'/{self.edt_xlsx_name.text()}_folder'
                        os.mkdir(self.loc_xlsx)

                        df_sheet = pd.DataFrame()
                        self.set_pre_lbl()  # id 정렬

                        # 엑셀 생성
                        wb = openpyxl.Workbook()
                        self.new_xlsx = self.loc_xlsx + fr'/{self.edt_xlsx_name.text()}.xlsx'
                        wb.save(self.new_xlsx)

                        self.sheet2_value()  # sheet2 landmark name 설정

                        for i in range(len(self.id_list)):  # 환자 수 만큼 만들고 df 합침
                            name = self.id_list[i].split('/')  # 환자 번호

                            lbl_chunk = self.landmark_id_format_change(self.lbl_id, self.id_list[i])  # 4개 단위로 리스트 나눔 (id,x,y,z)
                            pre_chunk = self.landmark_id_format_change(self.pre_id, self.id_list[i])

                            if len(lbl_chunk) != len(pre_chunk):
                                logger.error('label, predict id landmark data does not matched id :' + name[0])
                                continue

                            df_lbl = self.id_dataframe(lbl_chunk)
                            df_pre = self.id_dataframe(pre_chunk)

                            result = df_lbl.sub(df_pre)  # 결과값 데이터 프레임 df-df2
                            result['Landmark_num'] = df_lbl['Landmark_num']  # result[landmark_num] = 0이되서 정렬된 df[landmark_num] 넣음

                            # landmark 번호만 따로 생성 후 정렬
                            df_landmark = pd.DataFrame(lbl_chunk, columns=['Landmark_num', 'x', 'y', 'z'])  # 랜드마크 번호, 이름에 대한 dataframe 생성 3D

                            df_landmark['Landmark_num'] = df_landmark['Landmark_num'].astype(int)
                            df_landmark.insert(1, 'Landmark_name', self.landmark_name)
                            self.drop_landmark(df_landmark)  # x,y,z 제거
                            df_landmark = df_landmark.sort_values(by='Landmark_num')  # 데이터 정렬

                            # aver dataframe 생성
                            new_df = pd.DataFrame({'Landmark_num': [''], 'Landmark_name': ['Aver']})

                            # 정렬한 df의 아래에 aver 붙임
                            df_landmark = pd.concat([df_landmark, new_df], ignore_index=True)

                            result[name[0]] = (result['x'].pow(2) + result['y'].pow(2) + result['z'].pow(2)).pow(
                                1 / 2)  # name[-2] 파일명 뒤에 있는 환자 번호, 두 점 사이의 거리 공식 적용 3D
                            self.drop_landmark(result)
                            result[name[0]].loc[-1] = result[name[0]].mean(axis=0)  # 평균 axis = 0 : 행방향, axis =1 : 열방향
                            list_land = result[name[0]].tolist()  # 다음 df에 넣기 위해 list 로 만듬

                            # 엑셀에 id 처음 입력
                            patient_id = name[0].split('.')[0]

                            df_sheet.insert(0, patient_id, list_land)  # 새로운 데이터 프레임 첫 번째에 추가됨. (0, 이름, 결과)

                        # outlier 의 수치 이하의 값만 출력 후 평균값 만들어서 landmark 와 합침
                        df_sheet_outlier = df_sheet[df_sheet < float(self.edt_outlier.text())]
                        aver_outlier = average(df_sheet_outlier)
                        df_sheet_outlier['Aver'] = df_sheet_outlier.mean(axis=1)
                        df_result_outlier = pd.concat([df_landmark, df_sheet_outlier], axis=1)

                        # 평균값 만들어서 landmark 와 합침
                        aver = average(df_sheet)
                        df_sheet['Aver'] = df_sheet.mean(axis=1)  # 마지막 열에 평균 추가
                        df_result = pd.concat([df_landmark, df_sheet], axis=1)  # 랜드마크, value 데이터 프레임 합치기

                        # self.group_num : json 에서 받은 data 의 그룹별 landmark list 1[N]형태
                        # sum 으로 하나의 list 로 만듬
                        group = sum(self.group_num, [])
                        self.number = [int(i.split('[')[0]) for i in group]
                        self.number.append('')

                        # json 에 있는 그룹 landmark 만 출력 하기 위해서 query 사용
                        # query 는 비교 연산자와 비슷하게 사용 즉 조건에 부합 하는 data 만 출력
                        df_result = df_result.query(f'Landmark_num == {self.number}')
                        df_result_outlier = df_result_outlier.query(f'Landmark_num == {self.number}')

                        # 기본값 세팅
                        # 측정값과 num,name 나누어 다시 평균 만들기 현재 평균은 group 을 제거한 값도 같이 평균한 값임
                        df_result1 = df_result.iloc[:, 0:2]
                        df_result2 = df_result.iloc[:, 2: len(df_result.columns)]
                        df_result2.drop(df_result2.tail(1).index)
                        df_result2.loc[df_lbl.shape[0]] = df_result2.mean(axis=0)

                        # outlier 세팅
                        # result1은 기본값과 같이 사용할 수 있음
                        df_result2_outlier = df_result_outlier.iloc[:, 2: len(df_result_outlier.columns)]
                        df_result2_outlier.drop(df_result_outlier.tail(1).index)
                        df_result2_outlier.loc[df_lbl.shape[0]] = df_result2_outlier.mean(axis=0)

                        # 기본
                        # 결측치에 -99999 입력 -> 엑셀에서 색상 변경시 숫자일 때만 가능 하기 때문, 마지막 행,열에 전체 aver 추가
                        # index 정렬
                        df_result_concat = pd.concat([df_result1, df_result2], axis=1)
                        df_result_concat.iat[-1, -1] = aver

                        # 기본 표준 편차 설정
                        df_copy = df_result_concat.copy()
                        df_copy2 = df_copy.iloc[:-1, 2:-1]
                        df_std_row = pd.DataFrame(df_copy2.std(axis=1, ddof=0))
                        df_std_row.columns = ['std']
                        # pandas 는 표준표준편차를 기준으로함, 모표준편차로 변경 (ddof=1 이 기본)
                        df_std_column = pd.DataFrame(df_copy2.std(ddof=0))
                        df_std_column = df_std_column.transpose()    # 행렬변환
                        df_std_column['Landmark_name'] = ['std']
                        df_std_column['Landmark_num'] = ['']
                        df_std_column.index = [-1]  # index 0 이되면 다른 곳에 추가로 값이 들어감
                        df_result_std = pd.concat([df_result_concat, df_std_column])
                        df_result_std = pd.concat([df_result_std, df_std_row], axis=1)

                        self.std = df_result_std['Aver'].head(-2).std(ddof=0)    # head(-2) Aver, std 제외한 표준편차

                        # outlier 세팅
                        df_result_outlier_concat = pd.concat([df_result1, df_result2_outlier], axis=1)
                        df_result_outlier_concat.iat[-1, -1] = aver_outlier

                        # outlier 표준 편차 설정
                        df_copy_outlier = df_result_outlier_concat.copy()
                        df_copy2_outlier = df_copy_outlier.iloc[:-1, 2:-1]
                        df_std_row_outlier = pd.DataFrame(df_copy2_outlier.std(axis=1, ddof=0))
                        df_std_row_outlier.columns = ['std']
                        df_std_column_outlier = pd.DataFrame(df_copy2_outlier.std(ddof=0))
                        df_std_column_outlier = df_std_column_outlier.transpose()    # 행 열 전환
                        df_std_column_outlier['Landmark_name'] = ['std']
                        df_std_column_outlier['Landmark_num'] = ['']
                        df_std_column_outlier.index = [-1]  # index 0 이되면 다른 곳에 추가로 값이 들어감

                        # 행열에 표준편차 합치기
                        df_result_std_outlier = pd.concat([df_result_outlier_concat, df_std_column_outlier])
                        df_result_std_outlier = pd.concat([df_result_std_outlier, df_std_row_outlier], axis=1)

                        self.std_outlier = df_result_std_outlier['Aver'].head(-2).std(ddof=0)    # head(-2) Aver, std 제외한 표준편차

                        aver_std = "Landmark_name == ['Aver','std']"
                        df_outlier_aver_std_row = df_result_std_outlier.query(aver_std)  # 표준 편차, 평균 row
                        df_outlier_aver_std_row = df_outlier_aver_std_row.replace(['Aver', 'std'], ['Remove_Outlier_Aver', 'Remove_Outlier_std'])
                        df_outlier_aver_std_column = df_result_std_outlier[['Aver', 'std']]  # 표준 편차, 평균 column
                        df_outlier_aver_std_column = df_outlier_aver_std_column.rename(columns={'Aver': 'Remove_Outlier_Aver', 'std': 'Remove_Outlier_std'})
                        df_result_std = pd.concat([df_result_std, df_outlier_aver_std_row])
                        df_result_std = pd.concat([df_result_std, df_outlier_aver_std_column], axis=1)

                        self.df_result = df_result_std.fillna(-99999)
                        self.df_result.reset_index(inplace=True, drop='index')

                        # 엑셀
                        writer = pd.ExcelWriter(self.new_xlsx, engine='openpyxl')
                        self.df_result.to_excel(writer, startcol=0, startrow=3,
                                                index=False, sheet_name='Sheet1')  # 0,3부터 엑셀로 저장, 인덱스 제거, Sheet1에 저장

                        self.sheet2(self.df_result, writer, aver, aver_outlier)
                        self.sheet1_setting(self.new_xlsx)
                        self.sheet2_setting(self.new_xlsx)
                        self.sheet3_setting(self.new_xlsx)

                        # error 출력
                        self.error_id()
                        messagebox('notice', 'Excel 생성이 완료 되었습니다.')
                    else:
                        messagebox('Warning', "동일한 파일명이 존재합니다. 다시 입력하세요")
                        logger.error("same file name exist")

                else:
                    pass

            else:
                messagebox('Warning', "파일명을 입력하세요")
                logger.error("no file name")

        # label, predict 선택 되지 않았을 때
        elif self.lbl_lbl.text() == '' or self.lbl_pre.text() == '':
            messagebox('Warning', "label 또는 predict 경로를 확인 하세요.")
            logger.error("label, predict location error")

        logger.info("btn_export out")

    # 시트 색상,테두리 설정
    def sheet1_setting(self, xlsx: str):
        logger.info('sheet1 start')
        wb = openpyxl.load_workbook(filename=xlsx)
        ws = wb['Sheet1']

        ws.cell(row=3, column=3).value = 'Patient_ID'
        ws.cell(row=3, column=3).font = Font(bold=True)
        ws.cell(3, 3).fill = self.blue_color
        ws.cell(4, 1).fill = self.blue_color
        ws.cell(4, 2).fill = self.blue_color

        # landmark_num, landmark_name 색상 5 = 시작하는 row, 3 = outlier, std 개수
        for j in range(ws.max_row - 5 - 3):
            ws.cell(5 + j, 1).border = self.thin_border
            ws.cell(5 + j, 1).fill = self.green_color
            ws.cell(5 + j, 2).border = self.thin_border
            ws.cell(5 + j, 2).fill = self.yellow_color

        # row Aver value 색상
        for row in range(5, ws.max_row + 1):
            for p in range(4):
                ws.cell(row=row, column=ws.max_column - p).fill = self.blue_color2
                ws.cell(row=row, column=ws.max_column - p).border = self.thin_border

        # col Aver value 색상
        for col in range(3, ws.max_column + 1):
            for w in range(4):
                ws.cell(row=ws.max_row - w, column=col).fill = self.blue_color2
                ws.cell(row=ws.max_row - w, column=col).border = self.thin_border

        # 수치에 따른 색상, 결측치 값,색상 변환
        for col in range(3, ws.max_column + 1):

            for row in range(5, ws.max_row + 1):

                data = float(ws.cell(row=row, column=col).value)

                if data > float(self.edt_error.text()):  # 특정 수치 이상 이면 색상 변함
                    if data > float(self.edt_outlier.text()):
                        ws.cell(row=row, column=col).fill = self.outlier_color
                        ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFF')
                        ws.cell(row=row, column=col).border = self.thin_border
                    else:
                        ws.cell(row=row, column=col).fill = self.yellow_color2
                        ws.cell(row=row, column=col).border = self.thin_border

                elif data == -99999:
                    ws.cell(row=row, column=col).value = ' '
                    ws.cell(row=row, column=col).fill = self.gray_color2
                    ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=4, column=col).fill = self.gray_color

        # landmark.dat 에 없는 값 색상 변환
        for row in range(5, ws.max_row):
            if ws.cell(row=row, column=2).value == 'None':
                ws.cell(row=row, column=2).fill = self.red_color

        # Aver 색상 변경
        for f in range(4):
            ws.cell(4, ws.max_column - f).fill = self.blue_color
            ws.cell(4, ws.max_column - f).border = self.thin_border
            ws.cell(ws.max_row - f, 2).fill = self.blue_color
            ws.cell(ws.max_row - f, 2).border = self.thin_border

        # table 에 작성된 값 삽입
        # 따로 작성 하는 것은 제일 마지막에 작성 해야 한다. 최대 row, column 을 인식 하기 때문
        ws.cell(row=1, column=3).value = f'Error Safe Zone : {self.edt_error.text()}{self.lbl_mm.text()}'
        ws.cell(row=1, column=3).font = Font(bold=True)
        ws.cell(row=1, column=6).value = f'Hyperparameter : Batch size = {self.table.item(0, 1).text()}' \
                                         f', Learning rate = {self.table.item(1, 1).text()}' \
                                         f', optimizer = {self.table.item(2, 1).text()}' \
                                         f', aug = {self.table.item(3, 1).text()} '
        ws.cell(row=1, column=6).font = Font(bold=True)

        # 셀 너비 설정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions[get_column_letter(ws.max_column)].width = 20
        ws.column_dimensions[get_column_letter(ws.max_column - 1)].width = 20

        # comment 에 작성된 값 삽입
        ws.cell(row=2, column=6).value = f'comment : {self.edt.toPlainText()}'
        ws.cell(row=2, column=6).font = Font(bold=True)

        ws.cell(row=2, column=3).value = f'outlier : {self.edt_outlier.text()}{self.lbl_outlier_unit.text()}'
        ws.cell(row=2, column=3).font = Font(bold=True)

        # Patient_ID 위에 있는 셀 병합
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=ws.max_column - 4)    # -4 = 평균, 표준편차
        wb.save(filename=xlsx)
        logger.info('sheet1 end')

    # 시트2 landmark 2[sella]형태로 만듬
    def sheet2_value(self):
        logger.info('landmark2 naming start')

        group = self.open_json()

        group_key = list(group.keys())
        group_value = list(group.values())
        self.group_num = list(group.values())

        sheet2_group = []
        for i in range(len(group_key)):

            for j in range(len(group_value[i])):

                for k in range(len(self.landmark_name_value)):  # landmark_name_value = 2[Sella] 형태
                    name = self.landmark_name_value[k].split('[')

                    if str(group_value[i][j]) == name[0]:  # name[0] = 랜드마크 번호
                        group_value[i][j] = self.landmark_name_value[k]  # value 즉 num 가 2[Sella] 형태가 됨

        # group 에 정의 되지 않은 landmark [None] 붙이기
        for i in range(len(group_value)):

            for j in range(len(group_value[i])):

                if ']' in str(group_value[i][j]):
                    pass

                else:
                    group_value[i][j] = str(group_value[i][j]) + '[None]'

        # group_name, landmark_name 합치기
        for i in range(len(group_key)):
            sheet2_group.append(group_key[i])

            for j in range(len(group_value[i])):
                sheet2_group.append(group_value[i][j])

        self.df_sheet2_name = pd.DataFrame()
        self.df_sheet2_name.insert(0, 'Name', sheet2_group)
        logger.info('landmark2 naming end')

    # sheet2 기본값, outlier 에 따른 값 넣기
    def sheet2(self, df: pd.DataFrame, writer: pd.ExcelWriter, avr: float, outlier_aver: float):
        df_sheet2_name_aver = pd.DataFrame()
        df_sheet2_name_aver['Name'] = df['Landmark_num'].astype(str) + '[' + df['Landmark_name'] + ']'  # 2[Sella] 형식으로 dataframe 만듬
        df_sheet2_name_aver['Aver'] = df['Aver']
        df_sheet2_name_aver['std'] = df['std']
        df_sheet2_name_aver['Remove_Outlier_Aver'] = df['Remove_Outlier_Aver']
        df_sheet2_name_aver['Remove_Outlier_std'] = df['Remove_Outlier_std']
        df_sheet2_name_aver = df_sheet2_name_aver.drop(df_sheet2_name_aver.index[len(df_sheet2_name_aver) - 1])  # 마지막 줄 제거
        df_sheet2 = self.df_sheet2_name.merge(df_sheet2_name_aver, on='Name', how='left')  # 2[Sella] aver 형태로 합침 # 빈 칸 Nan 으로 합쳐짐

        df_sheet2.to_excel(writer, startcol=0, startrow=3, index=False, sheet_name='Sheet2')
        new_df = pd.DataFrame({'Name': ['Total_aver'], 'Aver': [avr], 'std': [self.std], 'Remove_Outlier_Aver': [outlier_aver], 'Remove_Outlier_std': [self.std_outlier]})
        df_sheet2 = pd.concat([new_df, df_sheet2])
        df_sheet2 = df_sheet2.fillna('None')
        df_sheet2.to_excel(writer, startcol=0, startrow=2, index=False, sheet_name='Sheet2')
        writer.save()  # Sheet2 저장

    # sheet2 xlsx
    def sheet2_setting(self, xlsx: str):
        logger.info('sheet2 start')
        group = self.open_json()

        group_key = list(group.keys())
        group_value = list(group.values())

        wb = openpyxl.load_workbook(filename=xlsx)
        ws = wb['Sheet2']

        # table 에 default 값 출력
        ws.cell(row=1, column=2).value = f'Error Safe Zone : {self.edt_error.text()}{self.lbl_mm.text()}'
        ws.cell(row=1, column=6).value = f'Hyperparameter Batch size = {self.table.item(0, 1).text()}' \
                                         f', Learning rate = {self.table.item(1, 1).text()}' \
                                         f', optimizer = {self.table.item(2, 1).text()}' \
                                         f', aug = {self.table.item(3, 1).text()} '
        ws.cell(row=2, column=6).value = f'comment : {self.edt.toPlainText()}'

        ws.cell(row=1, column=2).font = Font(bold=True)
        ws.cell(row=1, column=6).font = Font(bold=True)
        ws.cell(row=2, column=6).font = Font(bold=True)

        ws.cell(row=2, column=2).value = f'outlier : {self.edt_outlier.text()}{self.lbl_outlier_unit.text()}'
        ws.cell(row=2, column=2).font = Font(bold=True)

        # num[landmark_name],aver 의 색상
        for row in range(5, ws.max_row + 1):
            for a in range(4):
                data = ws.cell(row=row, column=2 + a).value

                if data == -99999:
                    ws.cell(row=row, column=2 + a).value = ' '
                    ws.cell(row=row, column=2 + a).fill = self.gray_color2
                    ws.cell(row=row, column=2 + a).border = self.thin_border

                elif data == 'None':
                    ws.cell(row=row, column=2 + a).fill = self.orange_color
                    ws.cell(row=row, column=2 + a).border = self.thin_border

            ws.cell(row=row, column=1).fill = self.yellow_color
            ws.cell(row=row, column=1).border = self.thin_border

        # Total_aver, aver 색상
        for e in range(5):
            ws.cell(row=4, column=1 + e).fill = self.blue_color
            ws.cell(row=4, column=1 + e).border = self.thin_border

        group_row = 5  # 시작줄
        landmark_row = 6

        # group 색상 초록색
        for i in range(len(group_key)):
            # Group 색상
            for h in range(5):
                ws.cell(row=group_row, column=1 + h).fill = self.green_color
                ws.cell(row=group_row, column=1 + h).border = self.thin_border
                ws.cell(row=group_row, column=1 + h).font = Font(bold=True)

            # Group 평균, std
            for j in range(len(group_value)):
                num = 0
                num_outlier = 0
                result = 0  # 합
                result_outlier = 0
                dispersion = 0  # 분산
                dispersion_outlier = 0  # 분산
                # None, ' ' 을 제외한 합

                for row in range(landmark_row, landmark_row + len(group_value[i])):    # 합
                    data = ws.cell(row=row, column=2).value
                    if data == 'None' or data == ' ':
                        pass
                    else:
                        result += ws.cell(row=row, column=2).value
                        result_outlier += ws.cell(row=row, column=4).value
                        num_outlier += 1
                        num += 1

                aver = result / num
                outlier_aver = result_outlier / num_outlier

                for dis in range(landmark_row, landmark_row + len(group_value[i])):    # 분산
                    std_data = ws.cell(row=dis, column=2).value
                    if std_data == 'None' or std_data == ' ':
                        pass
                    else:
                        dispersion_outlier += ((ws.cell(row=dis, column=4).value - outlier_aver) ** 2) / num_outlier
                        dispersion += ((ws.cell(row=dis, column=2).value - aver) ** 2) / num

                ws[f'B{group_row}'] = aver  # Group 평균 삽입
                ws[f'D{group_row}'] = outlier_aver
                ws[f'C{group_row}'] = sqrt(dispersion)    # 표준편차 (루트 분산)
                ws[f'E{group_row}'] = sqrt(dispersion_outlier)

            group_row += len(group_value[i]) + 1
            landmark_row += len(group_value[i]) + 1

        for row in range(4, ws.max_row + 1):  # 결측치 제외 특정 수치 이상 빨강색
            for b in range(4):
                data = ws.cell(row=row, column=2 + b).value

                if data == 'None' or data == ' ':  # None, ' ' 은 str 이라 제외
                    pass

                elif float(data) > float(self.edt_error.text()):
                    if float(data) > float(self.edt_outlier.text()):
                        ws.cell(row=row, column=2 + b).fill = self.outlier_color
                        ws.cell(row=row, column=2 + b).font = Font(bold=True, color='FFFFFF')
                        ws.cell(row=row, column=2 + b).border = self.thin_border
                    else:
                        ws.cell(row=row, column=2 + b).fill = self.yellow_color2
                        ws.cell(row=row, column=2 + b).border = self.thin_border

        group_land_row = 5
        count = 0
        for h in range(len(group_value)):  # Group 바뀔때 마다 줄 띄우기
            ws.insert_rows(group_land_row + count)
            ws.insert_rows(group_land_row + count + 1)
            group_land_row += 1
            count += len(group_value[h]) + 2

        ws.column_dimensions['A'].width = 23
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        # 색상 정보
        ws['F3'].fill = self.orange_color
        ws['F3'] = 'None'
        ws['G3'] = 'id에 landmark가 없음'
        ws['J3'].fill = self.gray_color2
        ws['k3'] = '결측치'
        ws['F5'].fill = self.blue_color
        ws['F5'] = 'Total_aver'
        ws['H5'].fill = self.green_color
        ws['H5'] = 'Group_Name'
        ws['J5'].fill = self.yellow_color
        ws['J5'] = 'Landmark'
        ws['L5'].fill = self.red_color
        ws['L5'] = 'Error'

        wb.save(filename=xlsx)
        # 그래프가 없는 시트2 불러옴 시트에서 평균값을 만들어서 다시 불러와야함
        df = pd.read_excel(xlsx, sheet_name='Sheet2', header=2, usecols=[0, 1, 2, 3, 4])
        df = df.replace(to_replace=' ', value=-0.0001)  # ' '결측치 -> -0.0001    오차값은 음수가 없어서 음수값으로 결측치와 존재하지 않은것을 확인한다
        df = df.replace(to_replace='None', value=-0.0002)  # 아에 없는거 -> -0.0002
        df = df.dropna(axis=0)  # 줄 띄워서 생긴 결측치 제거
        df = round(df, 4)  # 소수 자리수
        self.graph(df, xlsx)
        logger.info('sheet2 end')

    # 그래프 생성
    def graph(self, df: pd.DataFrame, xlsx: str):
        logger.info('graph start')
        # 이미지 폴더 이름, 생성
        loc = xlsx.split('.')
        loc = loc[0].split('/')
        location = self.loc_xlsx + f'/{loc[-1]}_image'
        os.mkdir(location)

        group = self.open_json()

        key = list(group.keys())
        value = list(group.values())
        group_list = []

        # 그룹별 landmark 개수 리스트
        for i in range(len(key)):
            group_list.append(1 + len(value[i]))

        graph = df  # 시트2 데이터 프레임
        graph_dict = graph.to_dict('list')
        graph_value = list(graph_dict.values())

        start_row = 0
        image_insert = 7  # 이미지 삽입 시작 셀

        # 이미지를 엑셀에 넣기 위함
        wb = openpyxl.load_workbook(filename=xlsx)
        ws = wb['Sheet2']

        # total_aver 이름, 측정값 추가
        group_total_name = []
        group_total_value = []
        group_total_value_outlier = []
        group_total_name.append(graph_value[0][0])
        group_total_value.append(graph_value[1][0])
        group_total_value_outlier.append(graph_value[3][0])

        # 그룹 개수 만큼 그래프 생성
        for j in range(len(group_list)):
            # group_value[0][0] - total_aver 이라 [1]부터 해야됨
            # Total_aver, group 으로 묶음
            group_total_name.append(graph_value[0][1 + start_row])
            group_total_value.append(graph_value[1][1 + start_row])
            group_total_value_outlier.append(graph_value[3][1 + start_row])

            # group, group 의 landmark 로 묶음
            group_name = graph_value[0][1 + start_row: 1 + group_list[j] + start_row]
            group_value = graph_value[1][1 + start_row: 1 + group_list[j] + start_row]
            group_value_outlier = graph_value[3][1 + start_row: 1 + group_list[j] + start_row]

            # group 만 묶기 위해 group landmark 개수를 더해서 group 의 시작 위치로 감
            start_row += group_list[j]

            self.vertical_graph(group_name, group_value, group_value_outlier, location)  # group, landmark 그래프 제작

            img = Image(location + f'/{group_name[0]}.png')  # 파일 저장
            img.width = 800  # 픽셀 단위 사이즈 변경
            img.height = 225
            ws.add_image(img, f'D{image_insert}')
            image_insert += group_list[j] + 2  # 2칸씩 띄워서 삽입

        # total 이랑 group graph
        self.horizon_graph(group_total_value, group_total_value_outlier, group_total_name, location)

        img = Image(location + f'/{group_total_name[0]}.png')  # 파일 불러옴
        img.width = 650  # 픽셀 단위 사이즈 변경
        img.height = 1000
        ws.add_image(img, 'N3')

        wb.save(filename=xlsx)
        logger.info('graph end')
        # 최대치 ----- 20 ~ 30
        # 소수점 3자리

    # 세로 graph 제작
    def vertical_graph(self, x: list, y: list, y_out: list, location: str):

        plt.figure(figsize=(13, 3))  # graph 사이즈
        plt.ylim([-3, 15])  # 범위
        plt.axhline(y=0, color='black')  # horizon y=0을 기준점 검정색 선을 그음
        plt.axhline(y=float(self.edt_error.text()), color='red', linestyle='--')  # horizon y=0을 기준점 검정색 선을 그음
        plt.xticks(fontsize=8, rotation=-5)

        # 처음 색상 결정
        if y[0] > float(self.edt_error.text()):
            colors = ['#FFCCCC']
        else:
            colors = ['#C1F0C1']  # 초록 #C1F0C1

            # 일정 수치 이상 색 변환
        for j in range(len(x) - 1):
            if float(y[j + 1]) >= float(self.edt_error.text()):
                colors.append('#FFCCCC')  # error 빨강
            else:
                colors.append('#FFFFB3')  # 기본 노랑 #FFFFB3

        if y_out[0] > float(self.edt_error.text()):
            colors_out = ['#FFCCCC']  # 파랑 #FFCCCC
        else:
            colors_out = ['#C1F0C1']  # 노랑 #b3d9ff

        for f in range(len(x) - 1):
            if y_out[f + 1] > float(self.edt_error.text()):
                colors_out.append('#FFCCCC')  # error 빨강
            else:
                colors_out.append('#FFFFB3')  # group 초록 #C1F0C1

        # colors 리스트 삽입
        sns.set_palette(sns.color_palette(colors))
        bar = sns.barplot(x=x, y=y, edgecolor='black', alpha=0.6, linestyle='dashed', linewidth=1.5, palette=colors)  # edge color 테두리 투명도
        sns.barplot(x=x, y=y_out, edgecolor='black', alpha=1, palette=colors_out)

        bar.set(title=x[0])

        # bar 아래 글씨
        count = 0
        for p in bar.patches:  # 바에 내용 추가
            height = p.get_height()
            if count < len(bar.patches) / 2:
                if height == -0.0001:    # 결측치
                    bar.text(p.get_x() + p.get_width() / 2, -2, 'N/A', ha='center', size=10, color='red')

                elif height == -0.0002:    # NONE
                    bar.text(p.get_x() + p.get_width() / 2, -2, 'Empty', ha='center', size=10, color='orange')

                else:
                    bar.text(p.get_x() + p.get_width() / 2, -1.1, height, ha='center', size=10)
            elif count >= len(bar.patches) / 2:
                if height == -0.0001:
                    bar.text(p.get_x() + p.get_width() / 2, -2, 'N/A', ha='center', size=10, color='red')

                elif height == -0.0002:
                    bar.text(p.get_x() + p.get_width() / 2., -2, 'Empty', ha='center', size=10, color='orange')

                else:
                    bar.text(p.get_x() + p.get_width() / 2, -2.4, f'({height})', ha='center', size=10)
            count += 1
        # 바에 내용 추가
        plt.savefig(location + f'/{x[0]}.png')  # save 랑 show 의 위치가 바뀌면 save 는 실행되지 않는다
        # plt.show() 바로 볼수 있음
        plt.close()

    #  가로 graph
    def horizon_graph(self, x: list, x_out: list, y: list, location: str):
        plt.figure(figsize=(12, 20))
        plt.xlim([-3, 15])  # 범위
        plt.axvline(x=0, color='black')  # vertical
        plt.axvline(x=float(self.edt_error.text()), color='red', linestyle='--')  # vertical
        plt.yticks(fontsize=12)
        plt.xticks(fontsize=12)

        # 처음 색상 결정
        if x[0] > float(self.edt_error.text()):
            colors = ['#FFCCCC']  # 파랑 #FFCCCC
        else:
            colors = ['#b3d9ff']  # 노랑 #b3d9ff

        for j in range(len(y) - 1):

            if x[j + 1] > float(self.edt_error.text()):
                colors.append('#FFCCCC')  # error 빨강

            else:
                colors.append('#C1F0C1')  # group 초록 #C1F0C1

        if x_out[0] > float(self.edt_error.text()):
            colors_out = ['#FFCCCC']  # 파랑 #FFCCCC
        else:
            colors_out = ['#b3d9ff']  # 노랑 #b3d9ff

        for m in range(len(y) - 1):

            if x_out[m + 1] > float(self.edt_error.text()):
                colors_out.append('#FFCCCC')  # error 빨강

            else:
                colors_out.append('#C1F0C1')  # group 초록 #C1F0C1

        # sns.set_palette(sns.color_palette([colors))
        bar = sns.barplot(x=x, y=y, edgecolor='black', alpha=0.6, linestyle='dashed', linewidth=1.5, palette=colors)
        sns.barplot(x=x_out, y=y, edgecolor='black', alpha=1, palette=colors_out)

        bar.set(title=y[0])
        count = 0
        for p in bar.patches:  # 바에 내용 추가
            width = p.get_width()
            if count < len(bar.patches) / 2:
                if width == -0.0001:    # 결측치
                    bar.text(-2, p.get_y() + p.get_height() / 2, 'N/A', ha='center', size=10, color='red')

                elif width == -0.0002:    # NONE
                    bar.text(-2, p.get_y() + p.get_height() / 2, 'Empty', ha='center', size=10, color='orange')

                else:
                    bar.text(-0.8, p.get_y() + p.get_height() / 2, width, ha='center', size=12)
            elif count >= len(bar.patches) / 2:
                if width == -0.0001:
                    bar.text(-2, p.get_y() + p.get_height() / 2, 'N/A', ha='center', size=10, color='red')

                elif width == -0.0002:
                    bar.text(-2, p.get_y() + p.get_height() / 2, 'Empty', ha='center', size=10, color='orange')

                else:
                    bar.text(-2.2, p.get_y() + p.get_height() / 2, f'({width})', ha='center', size=12)
            count += 1

        plt.savefig(location + f'/{y[0]}.png')  # save, show 의 위치가 바뀌면 save 는 실행되지 않는다, 파일 저장
        plt.close()

    def sheet3_setting(self, xlsx: str):
        logger.info('sheet3 start')
        wb = openpyxl.load_workbook(filename=xlsx)
        ws = wb['Sheet2']
        sheet3 = wb.copy_worksheet(ws)
        sheet3.title = 'Sheet3'
        ws.delete_cols(5)
        ws.delete_cols(3)
        wb.save(filename=xlsx)

        df = pd.read_excel(xlsx, sheet_name='Sheet3', header=2, usecols=[0, 1, 2, 3, 4])
        df = df.replace(to_replace=' ', value=-0.0001)  # ' '결측치 -> -0.0001    오차값은 음수가 없어서 음수값으로 결측치와 존재하지 않은것을 확인한다
        df = df.replace(to_replace='None', value=-0.0002)  # 아에 없는거 -> -0.0002
        df = df.dropna(axis=0)  # 줄 띄워서 생긴 결측치 제거
        df = round(df, 4)  # 소수 자리수
        self.std_graph(df, xlsx)
        logger.info('sheet3 end')

    def std_graph(self, df: pd.DataFrame, xlsx: str):    # 표준편차 그래프
        logger.info('std graph start')
        # 이미지 폴더 이름, 생성
        loc = xlsx.split('.')
        loc = loc[0].split('/')
        location_outlier = self.loc_xlsx + f'/Remove_Outlier_std_{loc[-1]}_image'
        location = self.loc_xlsx + f'/std_{loc[-1]}_image'
        os.mkdir(location)
        os.mkdir(location_outlier)

        group = self.open_json()

        key = list(group.keys())
        value = list(group.values())
        group_list = []

        # 그룹별 landmark 개수 리스트
        for i in range(len(key)):
            group_list.append(1 + len(value[i]))

        graph = df  # 시트2 데이터 프레임
        graph_dict = graph.to_dict('list')
        graph_value = list(graph_dict.values())

        start_row = 0
        image_insert = 7  # 이미지 삽입 시작 셀

        # 이미지를 엑셀에 넣기 위함
        wb = openpyxl.load_workbook(filename=xlsx)
        ws = wb['Sheet3']

        # total_aver 이름, 측정값, 표준편차, outlier ~
        group_total_name = []
        group_total_value = []
        group_total_std = []
        group_total_value_outlier = []
        group_total_std_outlier = []

        group_total_name.append(graph_value[0][0])
        group_total_value.append(graph_value[1][0])
        group_total_std.append(graph_value[2][0])
        group_total_value_outlier.append(graph_value[3][0])
        group_total_std_outlier.append(graph_value[4][0])

        # 그룹 개수 만큼 그래프 생성
        for j in range(len(group_list)):
            # group_value[0][0] - total_aver 이라 [1]부터 해야됨
            # Total_aver, group 으로 묶음
            group_total_name.append(graph_value[0][1 + start_row])
            group_total_value.append(graph_value[1][1 + start_row])
            group_total_std.append(graph_value[2][1 + start_row])
            group_total_value_outlier.append(graph_value[3][1 + start_row])
            group_total_std_outlier.append(graph_value[4][1 + start_row])

            # group, group 의 landmark 로 묶음
            group_name = graph_value[0][1 + start_row: 1 + group_list[j] + start_row]
            group_value = graph_value[1][1 + start_row: 1 + group_list[j] + start_row]
            group_std = graph_value[2][1 + start_row: 1 + group_list[j] + start_row]
            group_value_outlier = graph_value[3][1 + start_row: 1 + group_list[j] + start_row]
            group_std_outlier = graph_value[4][1 + start_row: 1 + group_list[j] + start_row]

            # group 만 묶기 위해 group landmark 개수를 더해서 group 의 시작 위치로 감
            start_row += group_list[j]

            group_std_arr = np.array(group_std)    # 표준편차 그래프에 넣기 위해 편하게 numpy 사용해서 list/2
            self.std_vertical_graph(group_name, group_value, location, group_std_arr/2, 'std_')  # group, landmark 그래프 제작

            img = Image(location + f'/std_{group_name[0]}.png')  # 파일 불러옴
            img.width = 800  # 픽셀 단위 사이즈 변경
            img.height = 225
            ws.add_image(img, f'F{image_insert}')

            group_std_outlier_arr = np.array(group_std_outlier)
            self.std_vertical_graph(group_name, group_value_outlier, location_outlier, group_std_outlier_arr/2, 'std_Remove_Outlier_')  # outlier group, landmark 그래프 제작

            img_std = Image(location_outlier + f'/std_Remove_Outlier_{group_name[0]}.png')  # 파일 불러옴
            img_std.width = 800  # 픽셀 단위 사이즈 변경
            img_std.height = 225
            ws.add_image(img_std, f'Q{image_insert}')

            image_insert += group_list[j] + 2  # 2칸씩 띄워서 삽입

        group_total_std_arr = np.array(group_total_std)
        # total 이랑 group graph
        self.std_horizon_graph(group_total_value, group_total_name, location, group_total_std_arr/2, 'std_')    # 기본 그래프

        img = Image(location + f'/std_{group_total_name[0]}.png')  # 파일 불러옴
        img.width = 650  # 픽셀 단위 사이즈 변경
        img.height = 1000
        ws.add_image(img, 'AB3')

        group_total_std_outlier_arr = np.array(group_total_std_outlier)
        self.std_horizon_graph(group_total_value_outlier, group_total_name, location_outlier, group_total_std_outlier_arr/2, 'std_Remove_Outlier_')    # outlier 그래프

        img_std = Image(location_outlier + f'/std_Remove_Outlier_{group_total_name[0]}.png')  # 파일 불러옴
        img_std.width = 650  # 픽셀 단위 사이즈 변경
        img_std.height = 1000
        ws.add_image(img_std, 'AB49')

        wb.save(filename=xlsx)
        logger.info('std graph end')
        # 최대치 ----- 20 ~ 30
        # 소수점 3자리

    def std_vertical_graph(self, x: list, y: list, location: str, std: list, title: str):    # 표준편차 수직 그래프

        plt.figure(figsize=(13, 3))  # graph 사이즈
        plt.ylim([-3, 15])  # 범위
        plt.axhline(y=0, color='black')
        plt.axhline(y=float(self.edt_error.text()), color='red', linestyle='--')  # horizon y=0을 기준점 검정색 선을 그음
        plt.errorbar(x, y, yerr=std, color='black', ecolor='blue', fmt='.', alpha=0.5, elinewidth=2)
        plt.xticks(fontsize=8, rotation=-5)

        # 처음 색상 결정
        if y[0] > float(self.edt_error.text()):
            colors = ['#FFCCCC']
        else:
            colors = ['#C1F0C1']  # 초록 #C1F0C1

            # 일정 수치 이상 색 변환
        for j in range(len(x) - 1):
            if float(y[j + 1]) >= float(self.edt_error.text()):
                colors.append('#FFCCCC')  # error 빨강
            else:
                colors.append('#FFFFB3')  # 기본 노랑 #FFFFB3

        # colors 리스트 삽입
        sns.set_palette(sns.color_palette(colors))
        bar = sns.barplot(x=x, y=y, edgecolor='black', alpha=0.6, linewidth=1.5, palette=colors)  # edge color 테두리 투명도

        bar.set(title=f'{title}{x[0]}')

        # bar 아래 글씨
        count = 0
        for p in bar.patches:  # 바에 내용 추가
            height = p.get_height()

            if height == -0.0001:    # 결측치
                bar.text(p.get_x() + p.get_width() / 2, -2, 'N/A', ha='center', size=10, color='red')
            elif height == -0.0002:    # NONE
                bar.text(p.get_x() + p.get_width() / 2, -2, 'Empty', ha='center', size=10, color='orange')
            else:
                bar.text(p.get_x() + p.get_width() / 2, -2, height, ha='center', size=10)

            count += 1
        # 바에 내용 추가
        plt.savefig(location + f'/{title}{x[0]}.png')  # save 랑 show 의 위치가 바뀌면 save 는 실행되지 않는다
        # plt.show() 바로 볼수 있음
        plt.close()

    def std_horizon_graph(self, x: list, y: list, location: str, std: list, title: str):    # 표준편차 수평 그래프
        plt.figure(figsize=(12, 20))
        plt.xlim([-3, 15])  # 범위
        plt.axvline(x=0, color='black')
        plt.axvline(x=float(self.edt_error.text()), color='red', linestyle='--')
        plt.errorbar(x, y, xerr=std, color='black', ecolor='blue', fmt='.', alpha=0.7, elinewidth=2)
        plt.yticks(fontsize=12)
        plt.xticks(fontsize=12)

        # 처음 색상 결정
        if x[0] > float(self.edt_error.text()):
            colors = ['#FFCCCC']  # 파랑 #FFCCCC
        else:
            colors = ['#b3d9ff']  # 노랑 #b3d9ff

        for j in range(len(y) - 1):

            if x[j + 1] > float(self.edt_error.text()):
                colors.append('#FFCCCC')  # error 빨강

            else:
                colors.append('#C1F0C1')  # group 초록 #C1F0C1

        # sns.set_palette(sns.color_palette([colors))
        bar = sns.barplot(x=x, y=y, edgecolor='black', alpha=0.6, linewidth=1.5, palette=colors)
        bar.set(title=f'{title}{y[0]}')
        count = 0
        for p in bar.patches:  # 바에 내용 추가
            width = p.get_width()
            if width == -0.0001:
                bar.text(-2, p.get_y() + p.get_height() / 2, 'N/A', ha='center', size=10, color='red')

            elif width == -0.0002:
                bar.text(-2, p.get_y() + p.get_height() / 2, 'Empty', ha='center', size=10, color='orange')

            else:
                bar.text(-2, p.get_y() + p.get_height() / 2, width, ha='center', size=12)

            count += 1

        plt.savefig(location + f'/{title}{y[0]}.png')  # save, show 의 위치가 바뀌면 save 는 실행되지 않는다, 파일 저장
        plt.close()
