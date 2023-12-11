import os

import SimpleITK as sitk
import seaborn as sns

import numpy as np
import openpyxl
import pandas as pd
from PySide2.QtCore import Qt
from PySide2.QtGui import QDoubleValidator
from PySide2.QtWidgets import QWidget, QTableWidget, QTableWidgetItem, QPushButton, QLabel, QLineEdit, QPlainTextEdit, \
    QFileDialog, QDialog, QMessageBox, QRadioButton
from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, PatternFill, borders, Font, Alignment
from openpyxl.utils import get_column_letter

from tqdm import tqdm

from PreShin.loggers import logger

"""
    Volume, Polygon(dentition) Tooth Segmentation 성능 측정코드
    
    UI 의 가장 위에 Volume 과 Polygon 을 클릭하여 선택하는 버튼 있음
    
    Volume 은 nrrd파일로 구성되야함, Polygon은 txt로 구성됨.
    
    augmentation 파일에 대한 id 도 추가되어 있음.
"""

batch = '4'
rate = '2e-4'
optimizer = 'adam'
aug = '0'
comment = 'write comment'
safe_zone_error = '0.15'
outlier_error = '0.3'


def messagebox(text: str, i: str):
    """
    특정 작업시 메세지 박스 띄우는 함수
    :param text: 메인 제목 텍스트 ( ex: Warning, Notice, etc... )
    :param i: 메세지 박스 띄우는 이유 ( ex: 두 개의 파일의 개수가 일치하지 않습니다. )
    :return: None
    """
    signBox = QMessageBox()
    signBox.setWindowTitle(text)
    signBox.setText(i)

    signBox.setIcon(QMessageBox.Information)
    signBox.setStandardButtons(QMessageBox.Ok)
    signBox.exec_()


class Tooth_UI(QWidget):
    mode = 'diceloss'
    mode_volume_polygon = 'polygon'

    def __init__(self):
        super().__init__()
        self.dialog = QDialog()
        self.initUI()

    def initUI(self):
        self.table = QTableWidget(4, 2, self.dialog)
        self.table.setSortingEnabled(False)  # 정렬 기능
        self.table.resizeRowsToContents()
        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(0, 80)  # checkbox 열 폭 강제 조절.
        self.table.setColumnWidth(1, 80)

        self.table.setItem(0, 0, QTableWidgetItem('Batch size'))
        self.table.setItem(0, 1, QTableWidgetItem(batch))
        self.table.setItem(1, 0, QTableWidgetItem('Learning rate'))
        self.table.setItem(1, 1, QTableWidgetItem(rate))
        self.table.setItem(2, 0, QTableWidgetItem('Optimizer'))
        self.table.setItem(2, 1, QTableWidgetItem(optimizer))
        self.table.setItem(3, 0, QTableWidgetItem('Aug'))
        self.table.setItem(3, 1, QTableWidgetItem(aug))

        self.table.setHorizontalHeaderLabels(["Name", "Value"])
        self.table.setGeometry(20, 195, 180, 145)

        btn_pre_path = QPushButton(self.dialog)
        btn_lbl_path = QPushButton(self.dialog)
        btn_export = QPushButton(self.dialog)
        btn_manual = QPushButton(self.dialog)

        self.btn_polygon = QPushButton(self.dialog)
        self.btn_volume = QPushButton(self.dialog)

        btn_pre_path.setText('Predict Path')
        btn_lbl_path.setText('Label Path')
        btn_export.setText('Export Excel')
        btn_manual.setText('Open Manual')

        self.btn_polygon.setText('Polygon')
        self.btn_volume.setText('Volume')

        btn_manual.setGeometry(20, 10, 100, 20)
        btn_lbl_path.setGeometry(20, 35, 100, 20)
        btn_pre_path.setGeometry(20, 60, 100, 20)
        btn_export.setGeometry(220, 345, 120, 30)

        self.btn_polygon.setGeometry(140, 10, 100, 20)
        self.btn_volume.setGeometry(250, 10, 100, 20)

        btn_lbl_path.clicked.connect(self.btn_lbl_clicked)
        btn_pre_path.clicked.connect(self.btn_pre_clicked)
        btn_export.clicked.connect(self.btn_export_clicked)

        self.btn_polygon.clicked.connect(self.btn_polygon_clicked)
        self.btn_volume.clicked.connect(self.btn_volume_clicked)

        self.btn_volume.setEnabled(True)
        self.btn_polygon.setEnabled(False)

        self.edt_pre = QLineEdit(self.dialog)
        self.edt_lbl = QLineEdit(self.dialog)
        lbl_error = QLabel(self.dialog)
        lbl_outlier = QLabel(self.dialog)

        lbl_comment = QLabel(self.dialog)
        lbl_xlsx_name = QLabel(self.dialog)
        lbl_xlsx = QLabel(self.dialog)

        self.edt_error_rate = QLineEdit(self.dialog)
        self.edt_error_rate.setAlignment(Qt.AlignRight)
        self.edt_outlier_rate = QLineEdit(self.dialog)
        self.edt_outlier_rate.setAlignment(Qt.AlignRight)

        self.edt_xlsx_name = QLineEdit(self.dialog)
        self.edt_xlsx_name.setAlignment(Qt.AlignRight)  # 엑셀명

        self.radiobutton()

        self.edt_lbl.setGeometry(130, 35, 230, 20)
        self.edt_pre.setGeometry(130, 60, 230, 20)
        lbl_error.setGeometry(220, 240, 100, 20)
        lbl_outlier.setGeometry(220, 290, 100, 20)
        self.edt_error_rate.setGeometry(220, 260, 50, 20)  # 에러 입력
        self.edt_outlier_rate.setGeometry(220, 310, 50, 20)  # 아웃라이어 입력

        lbl_comment.move(20, 90)
        lbl_xlsx_name.move(20, 355)
        lbl_xlsx.move(173, 355)

        self.edt_xlsx_name.setGeometry(70, 350, 103, 20)

        lbl_error.setText('Error Safe Zone')
        lbl_outlier.setText('Remove Outlier')
        lbl_comment.setText('Comment')
        lbl_xlsx_name.setText('파일명 : ')
        lbl_xlsx.setText('.xlsx')

        # # 숫자만 입력, 범위 -> 왜 256까지가 아니라 999가 되는지 모르겠음
        # double_validator = QDoubleValidator(0.0, 99.0, 2, self)
        # double_validator.setNotation(QDoubleValidator.StandardNotation)
        # self.edt_error_rate.setValidator(double_validator)  # int 값만 입력가능
        # self.edt_outlier_rate.setValidator(double_validator)

        # 퍼센티지 에서 장수로 변환 했기 때문에 입력도 퍼센티지만 입력하게 함.
        # 장수를 퍼센테지로 변환 할 경우에는 반올림으로 인해서 값의 변동 폭이 있어 그래프가 일치하지 않을 경우가 생김
        self.edt_error_rate.setText(safe_zone_error)
        self.edt_outlier_rate.setText(outlier_error)

        self.edt = QPlainTextEdit(self.dialog)
        self.edt.setPlainText(comment)
        self.edt.setGeometry(20, 105, 340, 80)

        self.dialog.setWindowTitle('Tooth')
        self.dialog.setGeometry(500, 300, 370, 420)
        self.dialog.exec()

    def btn_polygon_clicked(self):
        self.btn_polygon.setEnabled(False)
        self.btn_volume.setEnabled(True)
        Tooth_UI.mode_volume_polygon = 'polygon'
        print('polygon')

    def btn_volume_clicked(self):
        self.btn_polygon.setEnabled(True)
        self.btn_volume.setEnabled(False)
        Tooth_UI.mode_volume_polygon = 'volume'
        print('volume')

    def radiobutton(self):
        self.rdbtn_dice = QRadioButton('dice', self.dialog)
        self.rdbtn_diceloss = QRadioButton('diceloss', self.dialog)
        self.rdbtn_iou = QRadioButton('iou', self.dialog)

        self.rdbtn_diceloss.setChecked(True)

        self.rdbtn_diceloss.clicked.connect(self.radiobutton_clicked)
        self.rdbtn_dice.clicked.connect(self.radiobutton_clicked)
        self.rdbtn_iou.clicked.connect(self.radiobutton_clicked)

        self.rdbtn_dice.setGeometry(220, 195, 120, 30)
        self.rdbtn_diceloss.setGeometry(270, 195, 120, 30)
        self.rdbtn_iou.setGeometry(220, 215, 120, 30)

    def radiobutton_clicked(self):
        if self.rdbtn_diceloss.isChecked():
            self.edt_error_rate.setText(safe_zone_error)
            self.edt_outlier_rate.setText(outlier_error)
            Tooth_UI.mode = 'diceloss'
            print('diceloss')
        elif self.rdbtn_dice.isChecked():
            self.edt_error_rate.setText(f'{round(1 - float(safe_zone_error), 2)}')
            self.edt_outlier_rate.setText(f'{round(1 - float(outlier_error), 2)}')
            Tooth_UI.mode = 'dice'

            print('dice')
        elif self.rdbtn_iou.isChecked():
            self.edt_error_rate.setText(f'{round(1 - float(safe_zone_error), 2)}')
            self.edt_outlier_rate.setText(f'{round(1 - float(outlier_error), 2)}')
            Tooth_UI.mode = 'iou'

            print('iou')

    # label 버튼 클릭 -> 디렉토리 입력
    def btn_lbl_clicked(self):
        logger.info('Label Button IN')
        loc = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(), QFileDialog.ShowDirsOnly)  # 창 title, 주소 나중에 변경

        # 폴더 경로 입력
        if loc != '':
            self.edt_lbl.setText(str(loc))
        else:
            self.edt_lbl.setText('')
        logger.info('Label Button OUT')

    # predict 버튼 클릭 -> 디렉토리 입력
    def btn_pre_clicked(self):
        logger.info('Predict Button IN')
        loc = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(), QFileDialog.ShowDirsOnly)  # 창 title, 주소 나중에 변경

        # 폴더 경로 입력
        if loc != '':
            self.edt_pre.setText(str(loc))
        else:
            self.edt_pre.setText('')
        logger.info('Predict Button OUT')

    # export 버튼 클릭 -> main event
    def btn_export_clicked(self):
        # label, predict 경로 없을때, 파일명 입력되지 않았을 때. 동일한 파일명 존재 할 때

        logger.info('Export Button IN')

        if self.edt_lbl.text() != '' and self.edt_pre.text() != '':  # label, predict 경로 입력

            try:  # lbl pre 폴더 경로 확인
                os.listdir(self.edt_pre.text())
                os.listdir(self.edt_lbl.text())
            except FileNotFoundError:
                messagebox('Warning', 'predict 또는 lbl 의 폴더 경로가 올바르지 않습니다.')
                logger.error("Folder path error")
            else:
                if self.edt_xlsx_name.text() != '':  # 파일명 입력 했을때
                    loc_xlsx = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(), QFileDialog.ShowDirsOnly)  # 선택한 경로 str 저장

                    if loc_xlsx != '':  # 폴더 선택 했을때
                        file = os.listdir(loc_xlsx)  # 엑셀 저장 위치에 있는 파일 읽기

                        if fr'{self.edt_xlsx_name.text()}' not in file:  # 동일한 파일명이 없을때
                            os.mkdir(f'{loc_xlsx}/{self.edt_xlsx_name.text()}')

                            tooth = Tooth()  # class 가져옴
                            # vol.pre_lbl_compare(self.edt_lbl.text(), self.edt_pre.text())  # lbl, pre 폴더에 존재 하는 폴더 목록 비교
                            dict_lbl = tooth.save_data_root(self.edt_lbl.text())  # label loc 읽기
                            dict_pre = tooth.save_data_root(self.edt_pre.text())  # predict loc 읽기

                            diceloss_dataframe = tooth.make_dice_loss_dataframe(dict_pre, dict_lbl)  # 데이터 불러오기

                            # data = {'LL1': 0.94, 'LL2': 1.0, 'LL3': 0.7, 'LL4': 0.4, 'LL5': 0.3, 'LL6': 0.1, 'LL7': 0.12, 'LL8': 0.88,
                            #         'LU1': 1.0, 'LU2': 0.32, 'LU3': 0.5, 'LU4': 0.78, 'LU5': 1.0, 'LU6': 1.0, 'LU7': 1.0, 'LU8': 0.0,
                            #         'RL1': 0.6, 'RL2': 1.0, 'RL3': 0.5, 'RL4': 0.3, 'RL5': 0.0, 'RL6': 0.35, 'RL7': 1.0, 'RL8': 0.3,
                            #         'RU1': 0.3, 'RU2': 0.9, 'RU3': 1.0, 'RU4': 1.0, 'RU5': None, 'RU6': 0.08, 'RU7': 0.9, 'RU8': 1.0}
                            # temp1 = pd.DataFrame.from_dict(data=data, orient='index', columns=['2'])
                            # temp2 = pd.DataFrame.from_dict(data=data, orient='index', columns=['5']) * 0.5
                            # temp3 = pd.DataFrame.from_dict(data=data, orient='index', columns=['7']) * 0.3

                            # diceloss_dataframe = pd.concat([temp1, temp2], axis=1)
                            # diceloss_dataframe = pd.concat([diceloss_dataframe, temp3], axis=1)

                            # 엑셀 생성
                            tooth.to_xlsx(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx', diceloss_dataframe,
                                          error_outlier=float(self.edt_outlier_rate.text()), error_rate=float(self.edt_error_rate.text()))

                            # sheet1 설정
                            tooth.sheet1_xlsx_style(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx',
                                                    error_rate=float(self.edt_error_rate.text()),
                                                    error_outlier=float(self.edt_outlier_rate.text()))

                            # sheet2,3 설정
                            tooth.sheet2_xlsx_style(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx',
                                                    error_rate=float(self.edt_error_rate.text()),
                                                    error_outlier=float(self.edt_outlier_rate.text()), sheet_name='보고용')
                            # sheet2,3 설정
                            tooth.sheet2_xlsx_style(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx',
                                                    error_rate=float(self.edt_error_rate.text()),
                                                    error_outlier=float(self.edt_outlier_rate.text()), sheet_name='분석용')

                            # ui 에 있는 comment 삽입
                            self.insert_comment(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx', 'all_data')
                            self.insert_comment(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx', '보고용')
                            self.insert_comment(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx', '분석용')

                            messagebox('notice', 'Excel 생성이 완료 되었습니다.')

                        else:
                            messagebox('Warning', "동일한 파일명이 존재합니다. 다시 입력하세요")
                            logger.error("same file name exist")

                    else:
                        pass  # 폴더 선택 하지 않았을 때 pass

                else:
                    messagebox('Warning', "파일명을 입력하세요")
                    logger.error("no file name")

        else:
            messagebox('Warning', "label 또는 predict 경로를 확인 하세요.")
            logger.error("label, predict location error")

        logger.info("Export Button OUT")

    def insert_comment(self, loc: str, xlsx: str, sheet: str):  # comment 삽입
        wb = openpyxl.load_workbook(filename=f'{loc}/{xlsx}')
        ws = wb[sheet]

        # sheet1 에는 error rate, sheet, accuracy
        if sheet == '분석용' or sheet == '보고용':
            col = 0
            # table 에 default 값 출력
            ws.cell(row=1, column=6).value = f'Hyperparameter Batch size = {self.table.item(0, 1).text()}' \
                                             f', Learning rate = {self.table.item(1, 1).text()}' \
                                             f', optimizer = {self.table.item(2, 1).text()}' \
                                             f', aug = {self.table.item(3, 1).text()} '
            ws.cell(row=2, column=6).value = f'comment : {self.edt.toPlainText()}'
            ws.cell(row=4, column=6).value = f'mode : {Tooth_UI.mode}'
            ws.cell(row=1, column=6).font = Font(bold=True)
            ws.cell(row=2, column=6).font = Font(bold=True)
            ws.cell(row=4, column=6).font = Font(bold=True)
        else:
            col = 14

            # table 에 default 값 출력
            ws.cell(row=1, column=20).value = f'Hyperparameter Batch size = {self.table.item(0, 1).text()}' \
                                              f', Learning rate = {self.table.item(1, 1).text()}' \
                                              f', optimizer = {self.table.item(2, 1).text()}' \
                                              f', aug = {self.table.item(3, 1).text()} '
            ws.cell(row=2, column=20).value = f'comment : {self.edt.toPlainText()}'
            ws.cell(row=3, column=20).value = f'mode : {Tooth_UI.mode}'
            ws.cell(row=1, column=20).font = Font(bold=True)
            ws.cell(row=2, column=20).font = Font(bold=True)
            ws.cell(row=3, column=20).font = Font(bold=True)

        if Tooth_UI.mode == 'diceloss':
            range_safe_zone = '0.0'
            range_out_rate = '1.0'
        else:
            range_safe_zone = '1.0'
            range_out_rate = '0.0'

        ws.cell(row=1, column=col + 2).value = f'Safe Zone : {range_safe_zone} ~ {float(self.edt_error_rate.text())}'
        ws.cell(row=2, column=col + 2).value = f'Error Safe Zone : {float(self.edt_error_rate.text())} ~ {float(self.edt_outlier_rate.text())}'
        ws.cell(row=3, column=col + 2).value = f'Remove Outlier Rate : {float(self.edt_outlier_rate.text())} ~ {range_out_rate} '
        ws.cell(row=2, column=col + 1).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 색상 노랑
        ws.cell(row=3, column=col + 1).fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # 빨강
        ws.cell(row=1, column=col + 1).border = Border(left=borders.Side(style='thin'),
                                                       right=borders.Side(style='thin'),
                                                       top=borders.Side(style='thin'),
                                                       bottom=borders.Side(style='thin'))
        ws.cell(row=2, column=col + 1).border = Border(left=borders.Side(style='thin'),
                                                       right=borders.Side(style='thin'),
                                                       top=borders.Side(style='thin'),
                                                       bottom=borders.Side(style='thin'))
        ws.cell(row=3, column=col + 1).border = Border(left=borders.Side(style='thin'),
                                                       right=borders.Side(style='thin'),
                                                       top=borders.Side(style='thin'),
                                                       bottom=borders.Side(style='thin'))
        ws.cell(row=1, column=col + 2).font = Font(bold=True)  # 글씨 굵게
        ws.cell(row=2, column=col + 2).font = Font(bold=True)
        ws.cell(row=3, column=col + 2).font = Font(bold=True)

        wb.save(filename=f'{loc}/{xlsx}')


class Tooth:
    def __init__(self):
        super().__init__()
        self.get_mode = Tooth_UI.mode
        self.get_volume_polygon_mode = Tooth_UI.mode_volume_polygon
        self.operator = '>'
        self.reverse_operator = '<'
        self.count_error_data = []
        self.aver = float
        self.std = float
        self.out_aver = float
        self.out_std = float
        self.tooth_name = [
            'LL1', 'LL2', 'LL3', 'LL4', 'LL5', 'LL6', 'LL7', 'LL8',
            'LU1', 'LU2', 'LU3', 'LU4', 'LU5', 'LU6', 'LU7', 'LU8',
            'RL1', 'RL2', 'RL3', 'RL4', 'RL5', 'RL6', 'RL7', 'RL8',
            'RU1', 'RU2', 'RU3', 'RU4', 'RU5', 'RU6', 'RU7', 'RU8'
        ]
        self.accuracy_aver_std = pd.DataFrame()
        self.sheet_aver_std = pd.DataFrame()
        self.error_rate_aver_std = pd.DataFrame()

        self.thin_border = Border(left=borders.Side(style='thin'),
                                  right=borders.Side(style='thin'),
                                  top=borders.Side(style='thin'),
                                  bottom=borders.Side(style='thin'))
        self.blue_color = PatternFill(start_color='b3d9ff', end_color='b3d9ff', fill_type='solid')
        self.green_color = PatternFill(start_color='c1f0c1', end_color='c1f0c1', fill_type='solid')
        self.red_color = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
        self.yellow_color = PatternFill(start_color='ffffb3', end_color='ffffb3', fill_type='solid')
        self.yellow_color2 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.gray_color2 = PatternFill(start_color='e0e0eb', end_color='e0e0eb', fill_type='solid')
        self.gray_color = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
        self.blue_color2 = PatternFill(start_color='ccf5ff', end_color='ccf5ff', fill_type='solid')
        self.outlier_color = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        self.white_color = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    def compare(self, operator, a, b):
        result = False
        if operator == '<':
            result = a < b
        elif operator == '>':
            result = a > b

        return result

    def save_data_root(self, loc: str):  # 정답 기준
        """
        data 의 주소와 id, 치아 파일명을 dict 형태로 저장
        :param loc: label id loc
        :return: dict
        """
        logger.info('data id, root save in')
        all_data = {'loc': loc}  #
        all_data_list = os.listdir(loc)  # 환자 list

        for i in all_data_list:
            if '#' not in i:
                continue
            else:
                if '_' in i:
                    split_data_name = i.split('#')
                    only_id = split_data_name[0]
                    aug_num = split_data_name[1].split('_')[1].split('.')[0]  # augmentation 된 데이터 _number id 에 추가
                    pt_id = str(only_id) + '_' + str(aug_num)

                else:
                    split_data_name = i.split('#')
                    pt_id = split_data_name[0]

                if pt_id not in all_data:
                    all_data[pt_id] = []
                all_data[pt_id].append(i.split('.')[0])  # 확장자를 제외한 이름 추가

        logger.info('data id, root save out')
        return all_data

    def make_dice_loss_dataframe(self, predict_data: dict, label_data: dict):
        """
        predict, label 을 공식에 맞게 연산
        여기서 nrrd 또는 mha 확인
        :param predict_data: predict - key id, value tooth_name, loc 정보 저장
        :param label_data: label dataframe
        :return: pd.Dataframe
        """
        logger.info(f'make dice loss dataframe start')

        cal_result = pd.DataFrame(index=self.tooth_name)  # tooth name 정보가 column name 으로 저장된 빈 dataframe 생성
        print(f'Total Num ID = {len(list(label_data.keys())) - 1}')
        count = 0

        logger.info(f'dice, iou, diceloss calculate start')

        for key in list(label_data.keys()):
            data = {}
            if key != 'loc':
                count += 1
                print(f'Now : {count}, id : {key}')
                if len(label_data[key]) != len(predict_data[key]) or key not in predict_data:  # id에 해당되는 데이터가 맞지 않을 때 error // predict 에 없을때도 제외함.
                    logger.error(f"id : {key} data not equal")
                    continue
                for value in tqdm(label_data[key], desc='Num_tooth'):
                    if '#' not in value:
                        continue
                    elif value in predict_data[key]:
                        pre = rf'{predict_data["loc"]}/{value}'
                        lbl = rf'{label_data["loc"]}/{value}'
                        result = self.calculate(pre, lbl)
                        if '_' in value:
                            data[value.split('#')[1].split('_')[0]] = result
                        else:
                            data[value.split('#')[1]] = result

                id_result_diceloss = pd.DataFrame.from_dict(data=data, orient='index', columns=[key])

                cal_result = pd.concat([cal_result, id_result_diceloss], axis=1)
        logger.info(f'dice, iou, diceloss calculate end')
        logger.info(f'make dice loss dataframe end ')
        return cal_result

    def calculate(self, pre: str, lbl: str):
        """
        dice, diceloss, iou 계산
        polygon 과 volume 모드 두개의 공식이 다름.
        설정한 모드에 따라서 return 값을 변경함

        :param pre: predict data 주소
        :param lbl: label data 주소
        :return: diceloss, dice, iou 값
        """

        # volume mode
        if self.get_volume_polygon_mode == 'volume':

            lbl = lbl + '.nrrd'
            pre = pre + '.nrrd'

            reader = sitk.ImageFileReader()
            reader.SetImageIO('NrrdImageIO')
            reader.SetFileName(lbl)
            imgOrg: sitk.Image = reader.Execute()

            imgRCNp: np.ndarray = sitk.GetArrayFromImage(imgOrg)
            imgRCNp = imgRCNp.ravel()
            mfn = imgRCNp  # index 역할

            reader = sitk.ImageFileReader()
            reader.SetImageIO('NrrdImageIO')
            reader.SetFileName(pre)
            imgOrg: sitk.Image = reader.Execute()

            imgRCNp: np.ndarray = sitk.GetArrayFromImage(imgOrg)
            imgRCNp = imgRCNp.ravel()
            prn = imgRCNp  # index 역할

            mfn_count = 0
            prn_count = 0

            mfn_index = 0  # 마스크의 복셀 카운팅
            prn_index = 0  # 예측의 복셀 카운팅

            for mv in mfn:  # 넘파이로 변형된 배열을 한복셀씩 for문을 돌린다.
                mfn_index += 1  # 마스크의 복셀카운팅 인덱스가 1부터 시작.
                if mv > 0:
                    mfn_count += 1  # 마스크 중에 1인 복셀값을 카운팅.

            for pv in prn:  # 넘파이로 변형된 배열을 한복셀씩 for문을 돌린다.
                prn_index += 1  # 예측의 복셀카운팅 인덱스가 1부터 시작.
                if pv > 0:
                    prn_count += 1  # 예측 중에 1인 복셀값을 카운팅.

            intersection_count = 0
            union_count = 0  # 합집합

            comp_index = 0
            for mv in mfn:  # 넘파이로 변형된 배열을 한복셀씩 for문을 돌린다.
                if mv > 0 or prn[comp_index] > 0:  # 정답과 예측 복셀이 둘중 하나가 1인 경우
                    union_count += 1  # 합집합
                    if mv > 0 and prn[comp_index] > 0:  # 정답과 예측 복셀이 모두 1인 경우
                        intersection_count += 1  # 교집합 복셀을 1씩 늘린다.
                comp_index += 1

            dice = intersection_count * 2 / (mfn_count + prn_count)  # dice 계산  공식
            dice_loss = 1 - dice  # dice_loss 계산 공식
            iou = intersection_count / union_count  # iou 계산 공식

            # 현재 모드에 따라서 출력하는 값이 다름
            if self.get_mode == 'diceloss':
                self.operator = '>'
                self.reverse_operator = '<'
                return dice_loss
            elif self.get_mode == 'iou':
                self.operator = '<'
                self.reverse_operator = '>'
                return iou
            elif self.get_mode == 'dice':
                self.operator = '<'
                self.reverse_operator = '>'
                return dice

        # polygon mode
        else:
            lbl = lbl + '.onsdat'
            pre = pre + '.onsdat'

            lbl_count = 0  # lbl class 개수
            pre_count = 0  # pre class 개수
            intersection_count = 0  # 교집합
            union_count = 0  # 합집합

            # 두 개의 파일을 동시에 열기
            with open(lbl, 'r') as lbl_f, open(pre, 'r') as pre_f:
                lbl_lines = lbl_f.readlines()  # 파일1의 모든 라인 읽기
                pre_lines = pre_f.readlines()  # 파일2의 모든 라인 읽기

                # 두 파일의 라인 개수 확인
                if len(lbl_lines) != len(pre_lines):
                    print("경고: 두 파일의 라인 개수가 다릅니다.")
                    self.count_error_data.append(lbl)
                    return None  # None 으로 넘겨서 xlsx 표에 회색으로 남도록 함.

                else:
                    # 두 파일의 내용을 동시에 읽어 오기
                    for lbl_line, pre_line in zip(lbl_lines, pre_lines):
                        # 각 파일의 각 줄에 대해 원하는 작업 수행
                        lbl_line = lbl_line.strip()  # 공백 제거
                        pre_line = pre_line.strip()

                        # print(f'파일1: {lbl_line}, 파일2: {pre_line}')
                        # ex) 10,0 => 10 = 인덱스 정보(전체 index 행 index와 다름), 0 = 해당되는 인덱스의 mesh 유무
                        # test 일때는 index 정보가 없어서 ',' 가 없는 경우에는 그대로 진행 하도록 함
                        if ',' in lbl_line:
                            lbl_line = lbl_line.split(',')[1]
                        if ',' in pre_line:
                            pre_line = pre_line.split(',')[1]

                        lbl_count += int(lbl_line)  # lbl, pre 개수 각각 카운트
                        pre_count += int(pre_line)

                        if int(lbl_line) + int(pre_line) > 0:
                            union_count += 1  # 두개의 합이 0보다 크면 합집합에 카운트
                            if int(lbl_line) + int(pre_line) > 1:
                                intersection_count += 1  # 두개의 합이 1보다 크면 교집합에 카운트 (둘다 1이면 2가되서 교집합임)

                    # dice iou diceloss 계산 공식 적용 ( apply the dice, diceloss, iou formula )
                    dice = intersection_count * 2 / (lbl_count + pre_count)
                    iou = intersection_count / union_count
                    dice_loss = 1 - dice

                    if self.get_mode == 'diceloss':
                        self.operator = '>'
                        self.reverse_operator = '<'
                        return dice_loss
                    elif self.get_mode == 'iou':
                        self.operator = '<'
                        self.reverse_operator = '>'
                        return iou
                    elif self.get_mode == 'dice':
                        self.operator = '<'
                        self.reverse_operator = '>'
                        return dice

    def aver_std_process(self, diceloss: pd.DataFrame):
        diceloss_count = diceloss.count(axis=1)
        diceloss_count = diceloss_count.sum(axis=0)

        print('총 데이터 개수 :', diceloss_count)

        diceloss_sum = diceloss.sum(axis=1)
        diceloss_sum = diceloss_sum.sum(axis=0)  # 총 합
        print('총 합 :', diceloss_sum)

        average = diceloss_sum / diceloss_count

        print('총 평균 :', average)

        # 표준편차 공식 참고
        df_diceloss_std = diceloss.sub(average)  # data - 평균
        df_diceloss_std = df_diceloss_std.pow(2)  # 의 제곱
        df_diceloss_std = df_diceloss_std.sum(axis=1)
        df_diceloss_std = df_diceloss_std.sum(axis=0)
        diceloss_std = (df_diceloss_std / diceloss_count) ** (1 / 2)
        print('총 표준편차 :', diceloss_std)

        print(f'에러 데이터 : {self.count_error_data}')
        return average, diceloss_std

    def to_xlsx(self, loc: str, xlsx: str, result: pd.DataFrame, error_outlier: float, error_rate: float):  # 엑셀 생성, 결과값 삽입
        logger.info('make xlsx start')

        writer = pd.ExcelWriter(f'{loc}/{xlsx}', engine='openpyxl')  # pandas 엑셀 작성

        if self.reverse_operator == '<':
            outlier_result = result[result < error_outlier]
        else:
            outlier_result = result[result > error_outlier]

        self.aver, self.std = self.aver_std_process(result)
        self.out_aver, self.out_std = self.aver_std_process(outlier_result)
        # 전체 데이터 의 평균, 표준편자 데이터 생성
        aver_r = pd.DataFrame(data=result.mean(axis=0)).transpose()
        aver_r.index = ['Aver']

        std_r = pd.DataFrame(data=result.std(axis=0)).transpose()
        std_r.index = ['Std']

        out_aver_r = pd.DataFrame(data=outlier_result.mean(axis=0)).transpose()
        out_aver_r.index = ['Remove_Outlier_Aver']

        out_std_r = pd.DataFrame(data=outlier_result.std(axis=0)).transpose()
        out_std_r.index = ['Remove_Outlier_Std']

        row_aver_std = pd.concat([aver_r, std_r], axis=0)
        row_aver_std = pd.concat([row_aver_std, out_aver_r], axis=0)
        row_aver_std = pd.concat([row_aver_std, out_std_r], axis=0)

        aver_c = pd.DataFrame(data=result.mean(axis=1))
        aver_c.columns = ['Aver']

        std_c = pd.DataFrame(data=result.std(axis=1))
        std_c.columns = ['Std']

        out_aver_c = pd.DataFrame(data=outlier_result.mean(axis=1))
        out_aver_c.columns = ['Remove_Outlier_Aver']

        out_std_c = pd.DataFrame(data=outlier_result.std(axis=1))
        out_std_c.columns = ['Remove_Outlier_Std']

        column_aver_std = pd.concat([aver_c, std_c], axis=1)
        column_aver_std = pd.concat([column_aver_std, out_aver_c], axis=1)
        column_aver_std = pd.concat([column_aver_std, out_std_c], axis=1)
        all_data = pd.concat([result, row_aver_std], axis=0)

        all_data = pd.concat([all_data, aver_c], axis=1)
        all_data = pd.concat([all_data, std_c], axis=1)
        all_data = pd.concat([all_data, out_aver_c], axis=1)
        all_data = pd.concat([all_data, out_std_c], axis=1)

        all_data = all_data.fillna(-99999).round(5)
        column_aver_std = column_aver_std.fillna(-99999).round(5)
        # df 엑셀에 입력
        all_data.to_excel(writer, startcol=0, startrow=3, sheet_name='all_data')
        new_df = pd.DataFrame(
            {'Name': ['Total'], 'Aver': [self.aver], 'Std': [self.std], 'Remove_Outlier_Aver': [self.out_aver], 'Remove_Outlier_Std': [self.out_std]})
        column_aver_std.to_excel(writer, startcol=0, startrow=4, header=None, sheet_name='보고용')
        new_df.to_excel(writer, startcol=0, startrow=2, index=False, sheet_name='보고용')

        column_aver_std.to_excel(writer, startcol=0, startrow=4, header=None, sheet_name='분석용')
        new_df.to_excel(writer, startcol=0, startrow=2, index=False, sheet_name='분석용')

        writer.close()

        total = pd.DataFrame(
            {'Aver': [self.aver], 'Std': [self.std], 'Remove_Outlier_Aver': [self.out_aver], 'Remove_Outlier_Std': [self.out_std]})
        total.index = ['Total']
        df_for_std_graph = pd.concat([total, column_aver_std])

        df_for_std_graph = df_for_std_graph.replace(to_replace=-99999, value=-0.02).round(3)

        self.graph(df_for_std_graph, loc, xlsx, error_rate)
        self.std_graph(df_for_std_graph, loc, xlsx, error_rate)

        logger.info('make xlsx end')

    # sheet1 에 스타일 적용
    def sheet1_xlsx_style(self, loc: str, xlsx: str, error_outlier: float, error_rate: float):
        logger.info('Sheet1 Apply Style Start')

        wb = openpyxl.load_workbook(filename=f'{loc}/{xlsx}')
        ws = wb['all_data']

        ws.column_dimensions['A'].width = 23
        ws.column_dimensions[get_column_letter(ws.max_column)].width = 20
        ws.column_dimensions[get_column_letter(ws.max_column - 1)].width = 20

        ws.cell(row=3, column=2).value = 'Patient_ID'
        ws.cell(row=3, column=2).font = Font(bold=True)
        ws.cell(3, 2).fill = self.blue_color
        ws.cell(4, 1).fill = self.blue_color

        # Aver 색상 변경
        for f in range(4):
            ws.cell(4, ws.max_column - f).fill = self.blue_color
            ws.cell(4, ws.max_column - f).border = self.thin_border
            ws.cell(ws.max_row - f, 1).fill = self.blue_color
            ws.cell(ws.max_row - f, 1).border = self.thin_border

        # row Aver value 색상
        for row in range(5, ws.max_row + 1):
            for p in range(4):
                ws.cell(row=row, column=ws.max_column - p).fill = self.blue_color2
                ws.cell(row=row, column=ws.max_column - p).border = self.thin_border

        # col Aver value 색상
        for col in range(2, ws.max_column + 1):
            for w in range(4):
                ws.cell(row=ws.max_row - w, column=col).fill = self.blue_color2
                ws.cell(row=ws.max_row - w, column=col).border = self.thin_border

        # 수치에 따른 색상, 결측치 값,색상 변환
        for col in range(2, ws.max_column + 1):

            for row in range(5, ws.max_row + 1):

                data = float(ws.cell(row=row, column=col).value)

                if self.compare(self.operator, data, error_rate) and data != -99999:  # 특정 수치 보다 크면 이면 색상 변함
                    if 'Std' in ws.cell(row=row, column=1).value or 'Std' in str(ws.cell(row=4, column=col).value):
                        pass
                    elif self.compare(self.operator, data, error_outlier):
                        ws.cell(row=row, column=col).fill = self.outlier_color
                        ws.cell(row=row, column=col).border = self.thin_border
                    else:
                        ws.cell(row=row, column=col).fill = self.yellow_color2
                        ws.cell(row=row, column=col).border = self.thin_border

                elif data == -99999:
                    ws.cell(row=row, column=col).value = ' '
                    ws.cell(row=row, column=col).fill = self.gray_color2
                    ws.cell(row=row, column=col).border = self.thin_border

                else:
                    pass
                ws.cell(row=4, column=col).fill = self.gray_color

        # table 에 작성된 값 삽입
        # 따로 작성 하는 것은 제일 마지막에 작성 해야 한다. 최대 row, column 을 인식 하기 때문
        ws.insert_rows(1)
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=ws.max_column - 4)  # -4 = 평균, 표준편차

        ws['B1'] = '전체 데이터'
        ws['B2'] = '최종 결과'
        ws['D1'] = 'Aver'
        ws['F1'] = 'Std'
        ws['H1'] = 'Remove_outlier_Aver'
        ws['K1'] = 'Remove_outlier_Std'
        ws['D2'] = self.aver
        ws['F2'] = self.std
        ws['H2'] = self.out_aver
        ws['K2'] = self.out_std
        ws.merge_cells('B1:C1')
        ws.merge_cells('B2:C2')
        ws.merge_cells('D1:E1')
        ws.merge_cells('F1:G1')
        ws.merge_cells('H1:J1')
        ws.merge_cells('K1:M1')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:G2')
        ws.merge_cells('H2:J2')
        ws.merge_cells('K2:M2')

        ws['A5'] = 'Tooth Name'
        ws['A5'].font = Font(bold=True)
        ws['A5'].border = self.thin_border
        ws['A5'].alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 19

        for i in range(6, ws.max_row + 1 - 4):
            ws.cell(i, 1).fill = self.yellow_color

        for i in range(2):
            for j in range(1, 13):
                ws.cell(i + 1, j + 1).font = Font(bold=True)
                ws.cell(i + 1, j + 1).border = self.thin_border
                ws.cell(i + 1, j + 1).fill = self.yellow_color
                ws.cell(i + 1, j + 1).alignment = Alignment(horizontal='center')

        wb.save(filename=f'{loc}/{xlsx}')

        logger.info('Sheet1 Apply Style End')

    # sheet2 에 스타일 적용
    def sheet2_xlsx_style(self, loc: str, xlsx: str, error_outlier: float, error_rate: float, sheet_name: str):
        logger.info('Sheet2,3 Apply Style Start')

        wb = openpyxl.load_workbook(filename=f'{loc}/{xlsx}')
        ws = wb[sheet_name]
        ws.insert_rows(1)

        # Aver 색상 변경
        for f in range(4):
            ws.cell(5, 2 + f).fill = self.green_color
            ws.cell(5, 2 + f).border = self.thin_border
            ws.cell(5, 2 + f).font = Font(bold=True)
        ws.cell(4, 1).fill = self.blue_color
        ws.cell(4, 1).border = self.thin_border
        ws.cell(5, 1).fill = self.green_color
        ws.cell(5, 1).border = self.thin_border
        ws.cell(5, 1).font = Font(bold=True)
        ws.cell(5, 1).alignment = Alignment(horizontal='center')

        for i in range(6, ws.max_row + 1):
            ws.cell(i, 1).fill = self.yellow_color

        # 수치에 따른 색상, 결측치 값,색상 변환
        for col in range(2, ws.max_column + 1):
            for row in range(5, ws.max_row + 1):
                if ws.cell(row=row, column=col).value is None:
                    data = 0
                else:
                    data = float(ws.cell(row=row, column=col).value)
                if self.compare(self.operator, data, error_rate) and data != -99999:  # 특정 수치 보다 크면 이면 색상 변함
                    if row == 5 or 'Std' in str(ws.cell(row=4, column=col).value):
                        pass
                    elif self.compare(self.operator, data, error_outlier):
                        ws.cell(row=row, column=col).fill = self.outlier_color
                        ws.cell(row=row, column=col).border = self.thin_border
                    else:
                        ws.cell(row=row, column=col).fill = self.yellow_color2
                        ws.cell(row=row, column=col).border = self.thin_border
                elif data == -99999:
                    ws.cell(row=row, column=col).value = ' '
                    ws.cell(row=row, column=col).fill = self.gray_color2
                    ws.cell(row=row, column=col).border = self.thin_border
                else:
                    pass
                ws.cell(row=4, column=col).fill = self.gray_color

        ws.column_dimensions['A'].width = 19

        if sheet_name == '보고용':
            ws.delete_cols(5, 1)
            ws.delete_cols(3, 1)
            ws.column_dimensions['C'].width = 19
        else:
            ws.column_dimensions['D'].width = 19
            ws.column_dimensions['E'].width = 19
        ws['F3'] = '-----------'
        ws['G3'] = 'Error Safe Zone'
        ws['F3'].font = Font(color='FF0000', bold=True)
        ws['G3'].font = Font(bold=True)

        wb.save(filename=f'{loc}/{xlsx}')

        logger.info('Sheet2,3 Apply Style End')

    # 그래프 생성
    def graph(self, df: pd.DataFrame, save_path: str, xlsx: str, error_rate):
        logger.info('graph start')
        # 이미지 폴더 이름, 생성
        loc = xlsx.split('.')
        loc = loc[0].split('/')
        location = save_path + f'/{loc[-1]}_image'
        os.mkdir(location)

        graph = df  # 시트2 데이터 프레임
        graph_dict = graph.to_dict('list')
        graph_value = list(graph_dict.values())

        IMAGE_INSERT = 7  # 이미지 삽입 시작 셀
        # 이미지를 엑셀에 넣기 위함
        wb = openpyxl.load_workbook(filename=save_path + '/' + xlsx)
        ws = wb['보고용']
        # total_aver 이름, 측정값 추가

        self.tooth_name.insert(0, 'total')
        group_total_name = self.tooth_name
        group_total_value = graph_value[0]
        group_total_value_outlier = graph_value[2]

        # group 만 묶기 위해 group landmark 개수를 더해서 group 의 시작 위치로 감
        self.vertical_graph(group_total_name, group_total_value, group_total_value_outlier, location, error_rate)  # group, landmark 그래프 제작
        img = Image(location + f'/total.png')  # 파일 저장
        img.width = 1200  # 픽셀 단위 사이즈 변경
        img.height = 225
        ws.add_image(img, f'F{IMAGE_INSERT}')
        wb.save(filename=save_path + '/' + xlsx)
        logger.info('graph end')
        # 최대치 ----- 20 ~ 30
        # 소수점 3자리

    # 가로 graph 제작
    def vertical_graph(self, x: list, y: list, y_out: list, location: str, error_rate):
        plt.figure(figsize=(20, 3))  # graph 사이즈
        plt.ylim([-0.3, 1.05])  # 범위
        plt.axhline(y=0, color='black')  # horizon y=0을 기준점 검정색 선을 그음
        plt.axhline(y=error_rate, color='red', linestyle='--')  # horizon y=0을 기준점 검정색 선을 그음
        plt.xticks(fontsize=10, rotation=-5)
        # 처음 색상 결정
        if self.compare(self.operator, y[0], error_rate):
            colors = ['#FFCCCC']    # error 빨강
        else:
            colors = ['#C1F0C1']  # 초록 #C1F0C1
            # 일정 수치 이상 색 변환
        for j in range(len(x) - 1):
            if self.compare(self.operator, float(y[j+1]), error_rate):
                colors.append('#FFCCCC')  # error 빨강
            else:
                colors.append('#FFFFB3')  # 기본 노랑 #FFFFB3

        if self.compare(self.operator, y_out[0], error_rate):
            colors_out = ['#FFCCCC']  # error 빨강
        else:
            colors_out = ['#C1F0C1']  # 초록 #C1F0C1

        for f in range(len(x) - 1):
            if self.compare(self.operator, y_out[f + 1], error_rate):
                colors_out.append('#FFCCCC')  # error 빨강
            else:
                colors_out.append('#FFFFB3')  # 기본 노랑 #FFFFB3

        # colors 리스트 삽입// 순서에 따라서 나중에 생성되는 plot 의 색상은 보이지 않게 되어 dice 냐 dice loss 냐에 따라 순서변경
        if self.operator == '>':
            color_list = [colors, colors_out]
            alpha = [0.6, 1]
        else:
            color_list = [colors_out, colors]
            alpha = [1, 0.6]

        sns.set_palette(sns.color_palette(color_list[0]))
        bar = sns.barplot(x=x, y=y, edgecolor='black', alpha=alpha[0], linestyle='dashed', linewidth=1.5, palette=color_list[0])  # edge color 테두리 투명도
        sns.barplot(x=x, y=y_out, edgecolor='black', alpha=alpha[1], palette=color_list[1])
        bar.set(title=x[0])
        # bar 아래 글씨
        count = 0
        for p in bar.patches:  # 바에 내용 추가
            height = p.get_height()
            if count < len(bar.patches) / 2:
                if height == -0.01:  # 결측치
                    bar.text(p.get_x() + p.get_width() / 2, -0.2, 'N/A', ha='center', size=10, color='red')
                elif height == -0.02:  # NONE
                    bar.text(p.get_x() + p.get_width() / 2, -0.2, 'Empty', ha='center', size=10, color='orange')
                else:
                    bar.text(p.get_x() + p.get_width() / 2, -0.11, height, ha='center', size=10)
            elif count >= len(bar.patches) / 2:
                if height == -0.01:
                    bar.text(p.get_x() + p.get_width() / 2, -0.2, 'N/A', ha='center', size=10, color='red')
                elif height == -0.02:
                    bar.text(p.get_x() + p.get_width() / 2., -0.2, 'Empty', ha='center', size=10, color='orange')
                else:
                    bar.text(p.get_x() + p.get_width() / 2, -0.24, f'({height})', ha='center', size=10)
            count += 1
        # 바에 내용 추가
        plt.savefig(location + f'/{x[0]}.png')  # save 랑 show 의 위치가 바뀌면 save 는 실행되지 않는다
        # plt.show() 바로 볼수 있음
        plt.close()

    def std_graph(self, df: pd.DataFrame, save_path: str, xlsx: str, error_rate):  # 표준편차 그래프
        logger.info('std graph start')
        # 이미지 폴더 이름, 생성
        loc = xlsx.split('.')
        loc = loc[0].split('/')
        location_outlier = save_path + f'/Remove_Outlier_std_{loc[-1]}_image'
        location = save_path + f'/std_{loc[-1]}_image'
        os.mkdir(location)
        os.mkdir(location_outlier)

        graph = df  # 시트2 데이터 프레임
        graph_dict = graph.to_dict('list')
        graph_value = list(graph_dict.values())

        # 이미지를 엑셀에 넣기 위함
        wb = openpyxl.load_workbook(filename=save_path + '/' + xlsx)
        ws = wb['분석용']

        # total_aver 이름, 측정값, 표준편차, outlier
        group_total_name = self.tooth_name
        group_total_value = graph_value[0]
        group_total_value_outlier = graph_value[2]
        group_total_std = graph_value[1]
        group_total_std_outlier = graph_value[3]

        # group 만 묶기 위해 group landmark 개수를 더해서 group 의 시작 위치로 감
        group_std_arr = np.array(group_total_std)  # 표준편차 그래프에 넣기 위해 편하게 numpy 사용해서 list/2
        self.std_vertical_graph(group_total_name, group_total_value, location, group_std_arr / 2, 'std_', error_rate)  # group, landmark 그래프 제작
        img = Image(location + f'/std_total.png')  # 파일 불러옴
        img.width = 1200  # 픽셀 단위 사이즈 변경
        img.height = 225
        ws.add_image(img, f'F7')
        group_std_outlier_arr = np.array(group_total_std_outlier)
        # outlier group, landmark 그래프 제작
        self.std_vertical_graph(group_total_name, group_total_value_outlier, location_outlier, group_std_outlier_arr / 2, 'std_Remove_Outlier_', error_rate)
        img_std = Image(location_outlier + f'/std_Remove_Outlier_total.png')  # 파일 불러옴
        img_std.width = 1200  # 픽셀 단위 사이즈 변경
        img_std.height = 225
        ws.add_image(img_std, f'F18')
        wb.save(filename=save_path + '/' + xlsx)
        logger.info('std graph end')
        # 최대치 ----- 20 ~ 30
        # 소수점 3자리

    def std_vertical_graph(self, x: list, y: list, location: str, std: list, title: str, error_rate):  # 표준편차 수직 그래프

        plt.figure(figsize=(20, 3))  # graph 사이즈
        plt.ylim([-0.3, 1.05])  # 범위
        plt.axhline(y=0, color='black')
        plt.axhline(y=error_rate, color='red', linestyle='--')  # horizon y=0을 기준점 검정색 선을 그음
        plt.errorbar(x, y, yerr=std, color='black', ecolor='blue', fmt='.', alpha=0.5, elinewidth=2)
        plt.xticks(fontsize=10, rotation=-5)

        # 처음 색상 결정
        if self.compare(self.operator, y[0], error_rate):
            colors = ['#FFCCCC']    # error 빨강
        else:
            colors = ['#C1F0C1']  # 초록 #C1F0C1

            # 일정 수치 이상 색 변환
        for j in range(len(x) - 1):
            if self.compare(self.operator, y[j + 1], error_rate):
                colors.append('#FFCCCC')  # error 빨강
            else:
                colors.append('#FFFFB3')  # 기본 노랑 #FFFFB3

        # colors 리스트 삽입
        sns.set_palette(sns.color_palette(colors))
        bar = sns.barplot(x=x, y=y, edgecolor='black', alpha=0.6, linewidth=1.5, palette=colors)  # edge color 테두리 투명도

        bar.set(title=f'{title}{x[0]}')

        # bar 아래 글씨
        count = 0
        for p in bar.patches:  # 바에 내용 추가
            height = p.get_height()

            if height == -0.01:  # 결측치
                bar.text(p.get_x() + p.get_width() / 2, -0.2, 'N/A', ha='center', size=10, color='red')
            elif height == -0.02:  # NONE
                bar.text(p.get_x() + p.get_width() / 2, -0.2, 'Empty', ha='center', size=10, color='orange')
            else:
                bar.text(p.get_x() + p.get_width() / 2, -0.2, height, ha='center', size=10)

            count += 1
        # 바에 내용 추가
        plt.savefig(location + f'/{title}{x[0]}.png')  # save 랑 show 의 위치가 바뀌면 save 는 실행되지 않는다
        # plt.show() 바로 볼수 있음
        plt.close()

# if __name__ == "__main__":
#     vol = Vol_Template()
#     xlsx_name = 'result.xlsx'
#     df_lbl = vol.label(r'D:\temp_data\label')
#     df_pre = vol.predict(r'D:\temp_data\predict')
#
#     image_folder = vol.create_img_folder(r'D:\temp_data', 'ff')
#     vol.to_xlsx(fr'D:\temp_data', xlsx_name, df_lbl, df_pre, 5, 15)
#     vol.sheet1_xlsx_style(fr'D:\temp_data', xlsx_name, 5, 13, 3, 7, image_folder, 'ff')
#     vol.sheet2_xlsx_style(fr'D:\temp_data', xlsx_name, 5, 3, image_folder, 'ff')
# fr'D:\temp_data\graph_image'
