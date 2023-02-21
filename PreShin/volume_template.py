import os
import re

import numpy as np
import openpyxl
import pandas as pd
from PySide2.QtCore import Qt
from PySide2.QtGui import QDoubleValidator
from PySide2.QtWidgets import QWidget, QTableWidget, QTableWidgetItem, QPushButton, QLabel, QLineEdit, QPlainTextEdit, QFileDialog, QDialog, QMessageBox
from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, PatternFill, borders, Font, Alignment
import matplotlib.ticker as mticker
from PreShin.loggers import logger
'''
    모든 기준 값은 input 값을 장수로 표현한 것을 기준으로 오차, 평균 등을 구한다.
'''
batch = '4'
rate = '2e-4'
optimizer = 'adam'
aug = '0'
comment = 'write comment'
safe_zone_error = '2'
outlier_error = '4'

error_rate_range = 10    # 축 범위 변경.
sheet_range = 40
accuracy_range = 100
# 엑셀 필터 기능이 문서에는 적용이 안된다고 하는데 동작이 되서 개수가 많아질 경우 확인해야함.

# 필요한 predict 파일 : air, sts, hts 정보가 들어있는 predict txt 파일 폴더들
# 필요한 label 파일 : air, sts, hts png label 파일이 들어있는 폴더들
# predict 파일의 구조는 아직 id 당 각각 폴더 안에 있을지 하나의 폴더에 있을지 정해지지 않음

def messagebox(text: str, i: str):
    signBox = QMessageBox()
    signBox.setWindowTitle(text)
    signBox.setText(i)

    signBox.setIcon(QMessageBox.Information)
    signBox.setStandardButtons(QMessageBox.Ok)
    signBox.exec_()


class Vol_Template_UI(QWidget):
    def __init__(self):
        super().__init__()
        self.dialog = QDialog()
        self.initUI()

    def initUI(self):
        self.table = QTableWidget(4, 2, self.dialog)
        self.table.setSortingEnabled(False)  # 정렬 기능
        self.table.resizeRowsToContents()
        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(0, 80)  # checkbox 열 폭 강제 조절.
        self.table.setColumnWidth(1, 80)

        self.table.setItem(0, 0, QTableWidgetItem('Batch size'))
        self.table.setItem(0, 1, QTableWidgetItem(batch))
        self.table.setItem(1, 0, QTableWidgetItem('Learning rate'))
        self.table.setItem(1, 1, QTableWidgetItem(rate))
        self.table.setItem(2, 0, QTableWidgetItem('Optimizer'))
        self.table.setItem(2, 1, QTableWidgetItem(optimizer))
        self.table.setItem(3, 0, QTableWidgetItem('Aug'))
        self.table.setItem(3, 1, QTableWidgetItem(aug))

        self.table.setHorizontalHeaderLabels(["Name", "Value"])
        self.table.setGeometry(20, 195, 180, 145)

        btn_pre_path = QPushButton(self.dialog)
        btn_lbl_path = QPushButton(self.dialog)
        btn_export = QPushButton(self.dialog)
        btn_manual = QPushButton(self.dialog)

        btn_pre_path.setText('Predict Path')
        btn_lbl_path.setText('Label Path')
        btn_export.setText('Export Excel')
        btn_manual.setText('Open Manual')

        self.edt_pre = QLineEdit(self.dialog)
        self.edt_lbl = QLineEdit(self.dialog)
        lbl_error = QLabel(self.dialog)
        lbl_outlier = QLabel(self.dialog)

        lbl_error_rate = QLabel(self.dialog)
        lbl_error_sheet = QLabel(self.dialog)
        lbl_outlier_rate = QLabel(self.dialog)
        lbl_outlier_sheet = QLabel(self.dialog)

        lbl_comment = QLabel(self.dialog)
        lbl_xlsx_name = QLabel(self.dialog)
        lbl_xlsx = QLabel(self.dialog)

        self.lbl_error_sheet = QLabel(self.dialog)
        self.lbl_error_sheet.setAlignment(Qt.AlignRight)  # 우측 정렬
        self.edt_error_rate = QLineEdit(self.dialog)
        self.edt_error_rate.setAlignment(Qt.AlignRight)
        self.lbl_outlier_sheet = QLabel(self.dialog)
        self.lbl_outlier_sheet.setAlignment(Qt.AlignRight)
        self.edt_outlier_rate = QLineEdit(self.dialog)
        self.edt_outlier_rate.setAlignment(Qt.AlignRight)

        self.edt_xlsx_name = QLineEdit(self.dialog)
        self.edt_xlsx_name.setAlignment(Qt.AlignRight)  # 엑셀명

        self.edt_lbl.setGeometry(130, 35, 230, 20)
        self.edt_pre.setGeometry(130, 60, 230, 20)
        lbl_error.setGeometry(220, 200, 100, 20)
        lbl_outlier.setGeometry(220, 269, 100, 20)

        lbl_error_rate.setGeometry(275, 223, 100, 20)
        lbl_error_sheet.setGeometry(274, 246, 100, 20)
        lbl_outlier_rate.setGeometry(275, 292, 100, 20)
        lbl_outlier_sheet.setGeometry(274, 315, 100, 20)

        lbl_comment.move(20, 90)
        lbl_xlsx_name.move(20, 355)
        lbl_xlsx.move(173, 355)

        self.lbl_error_sheet.setGeometry(215, 250, 50, 20)
        self.edt_error_rate.setGeometry(220, 223, 50, 20)  # 퍼센티지 에러 입력
        self.lbl_outlier_sheet.setGeometry(215, 319, 50, 20)
        self.edt_outlier_rate.setGeometry(220, 292, 50, 20)  # 퍼센티지 아웃라이어 입력

        self.edt_xlsx_name.setGeometry(70, 350, 103, 20)

        lbl_error_rate.setText('%')
        lbl_error_sheet.setText('장')
        lbl_outlier_rate.setText('%')
        lbl_outlier_sheet.setText('장')

        lbl_error.setText('Error Safe Zone')
        lbl_outlier.setText('Remove Outlier')
        lbl_comment.setText('Comment')
        lbl_xlsx_name.setText('파일명 : ')
        lbl_xlsx.setText('.xlsx')

        # 숫자만 입력, 범위 -> 왜 256까지가 아니라 999가 되는지 모르겠음
        double_validator = QDoubleValidator(0.0, 99.0, 2, self)
        double_validator.setNotation(QDoubleValidator.StandardNotation)
        self.edt_error_rate.setValidator(double_validator)  # int 값만 입력가능
        self.edt_outlier_rate.setValidator(double_validator)

        # 퍼센티지 에서 장수로 변환 했기 때문에 입력도 퍼센티지만 입력하게 함.
        # 장수를 퍼센테지로 변환 할 경우에는 반올림으로 인해서 값의 변동 폭이 있어 그래프가 일치하지 않을 경우가 생김
        self.edt_error_rate.setText(safe_zone_error)
        self.edt_error_rate.textChanged[str].connect(self.lbl_error_changed)
        self.lbl_error_sheet.setText(str(round((int(self.edt_error_rate.text()) * 256 / 100))))

        self.edt_outlier_rate.setText(outlier_error)
        self.edt_outlier_rate.textChanged[str].connect(self.lbl_outlier_changed)
        self.lbl_outlier_sheet.setText(str(round((int(self.edt_outlier_rate.text()) * 256 / 100))))

        btn_manual.setGeometry(20, 10, 100, 20)
        btn_lbl_path.setGeometry(20, 35, 100, 20)
        btn_pre_path.setGeometry(20, 60, 100, 20)
        btn_export.setGeometry(220, 345, 120, 30)

        btn_lbl_path.clicked.connect(self.btn_lbl_clicked)
        btn_pre_path.clicked.connect(self.btn_pre_clicked)
        btn_export.clicked.connect(self.btn_export_clicked)
        # btn_manual.clicked.connect(btn_manual_clicked)

        self.edt = QPlainTextEdit(self.dialog)
        self.edt.setPlainText(comment)
        self.edt.setGeometry(20, 105, 340, 80)

        self.dialog.setWindowTitle('AI')
        self.dialog.setGeometry(500, 300, 370, 420)
        self.dialog.exec()

    # Qlineedit 에 입력된 error 를 label 에 % 로 변환하고 실시간 전환
    def lbl_error_changed(self, text):
        # 001, 0001, 01, 0, 00, ... , None 에 맞춰서 변환
        if text == '' or text == '00':
            self.edt_error_rate.setText('0')
            text = 0

        elif len(text) >= 2 and text.startswith('0') is True and '.' not in text:
            text = text[1:]
            self.edt_error_rate.setText(text)

        self.lbl_error_sheet.setText(str(round((float(text) * 256 / 100))))

    # Qlineedit 에 작성된 outlier 값을 변환
    def lbl_outlier_changed(self, text):
        if text == '' or text == '00':
            self.edt_outlier_rate.setText('0')
            text = 0

        elif len(text) >= 2 and text.startswith('0') is True and '.' not in text:
            text = text[1:]
            self.edt_outlier_rate.setText(text)

        self.lbl_outlier_sheet.setText(str(round((float(text) * 256 / 100))))

    # label 버튼 클릭 -> 디렉토리 입력
    def btn_lbl_clicked(self):
        logger.info('Label Button IN')
        loc = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(), QFileDialog.ShowDirsOnly)  # 창 title, 주소 나중에 변경

        # 폴더 경로 입력
        if loc != '':
            self.edt_lbl.setText(str(loc))
        else:
            self.edt_lbl.setText('')
        logger.info('Label Button OUT')

    # predict 버튼 클릭 -> 디렉토리 입력
    def btn_pre_clicked(self):
        logger.info('Predict Button IN')
        loc = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(), QFileDialog.ShowDirsOnly)  # 창 title, 주소 나중에 변경

        # 폴더 경로 입력
        if loc != '':
            self.edt_pre.setText(str(loc))
        else:
            self.edt_pre.setText('')
        logger.info('Predict Button OUT')

    # export 버튼 클릭 -> main event
    def btn_export_clicked(self):
        # label, predict 경로 없을때, 파일명 입력되지 않았을 때. 동일한 파일명 존재 할 때

        logger.info('Export Button IN')

        if self.edt_lbl.text() != '' and self.edt_pre.text() != '':  # label, predict 경로 입력

            try:  # lbl pre 폴더 경로 확인
                os.listdir(self.edt_pre.text())
                os.listdir(self.edt_lbl.text())
            except FileNotFoundError:
                messagebox('Warning', 'predict 또는 lbl 의 폴더 경로가 올바르지 않습니다.')
                logger.error(" 폴더 경로 에러")
            else:
                if self.edt_xlsx_name.text() != '':  # 파일명 입력 했을때
                    loc_xlsx = QFileDialog.getExistingDirectory(self, "Open file", os.getcwd(), QFileDialog.ShowDirsOnly)  # 선택한 경로 str 저장

                    if loc_xlsx != '':  # 폴더 선택 했을때
                        file = os.listdir(loc_xlsx)  # 엑셀 저장 위치에 있는 파일 읽기

                        if fr'{self.edt_xlsx_name.text()}' not in file:  # 동일한 파일명이 없을때
                            os.mkdir(f'{loc_xlsx}/{self.edt_xlsx_name.text()}')

                            vol = Vol_Template()  # class 가져옴
                            # vol.pre_lbl_compare(self.edt_lbl.text(), self.edt_pre.text())  # lbl, pre 폴더에 존재 하는 폴더 목록 비교
                            df_lbl = vol.label(self.edt_lbl.text())  # label loc 읽기
                            df_pre = vol.predict(self.edt_pre.text())  # predict loc 읽기

                            img_folder = vol.create_img_folder(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', self.edt_xlsx_name.text())  # image 폴더 생성

                            # 엑셀 생성
                            vol.to_xlsx(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx', df_lbl, df_pre,
                                        error_outlier=float(self.edt_outlier_rate.text()),
                                        sheet_outlier=int(self.lbl_outlier_sheet.text()))

                            # sheet1 설정
                            vol.sheet1_xlsx_style(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx',
                                                  error_rate=float(self.edt_error_rate.text()),
                                                  error_outlier=float(self.edt_outlier_rate.text()), sheet_outlier=int(self.lbl_outlier_sheet.text()),
                                                  sheet_error=int(self.lbl_error_sheet.text()),
                                                  image_folder_loc=img_folder, file_name=self.edt_xlsx_name.text())

                            # sheet2 설정
                            vol.sheet2_xlsx_style(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx',
                                                  error_rate=float(self.edt_error_rate.text()),
                                                  error_outlier=float(self.edt_outlier_rate.text()),
                                                  image_folder_loc=img_folder, file_name=self.edt_xlsx_name.text())

                            # ui 에 있는 comment 삽입
                            self.insert_comment(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx', '분석용')
                            self.insert_comment(f'{loc_xlsx}/{self.edt_xlsx_name.text()}', fr'{self.edt_xlsx_name.text()}.xlsx', '보고용')
                            messagebox('notice', 'Excel 생성이 완료 되었습니다.')

                        else:
                            messagebox('Warning', "동일한 파일명이 존재합니다. 다시 입력하세요")
                            logger.error("same file name exist")

                    else:
                        pass  # 폴더 선택 하지 않았을 때 pass

                else:
                    messagebox('Warning', "파일명을 입력하세요")
                    logger.error("no file name")

        else:
            messagebox('Warning', "label 또는 predict 경로를 확인 하세요.")
            logger.error("label, predict location error")

        logger.info("Export Button OUT")

    def insert_comment(self, loc: str, xlsx: str, sheet: str):  # comment 삽입
        wb = openpyxl.load_workbook(filename=f'{loc}/{xlsx}')
        ws = wb[sheet]
        # table 에 default 값 출력
        ws.cell(row=1, column=7).value = f'Hyperparameter Batch size = {self.table.item(0, 1).text()}' \
                                         f', Learning rate = {self.table.item(1, 1).text()}' \
                                         f', optimizer = {self.table.item(2, 1).text()}' \
                                         f', aug = {self.table.item(3, 1).text()} '
        ws.cell(row=2, column=7).value = f'comment : {self.edt.toPlainText()}'

        # sheet1 에는 error rate, sheet, accuracy
        if sheet == '분석용':
            ws.cell(row=1,
                    column=2).value = f'Safe Error rate : {self.edt_error_rate.text()} %  sheet : {self.lbl_error_sheet.text()} 장  (Accuracy : {100 - int(self.edt_error_rate.text())}%) '
            ws.cell(row=2,
                    column=2).value = f'Remove Outlier Rate : {self.edt_outlier_rate.text()} %  sheet : {self.lbl_outlier_sheet.text()} 장  (Accuracy : {100 - int(self.edt_outlier_rate.text())})%'

        elif sheet == '보고용':
            ws.cell(row=1, column=2).value = f'Error Safe Zone : {100 - float(self.edt_error_rate.text())} %  '
            ws.cell(row=2, column=2).value = f'Remove Outlier Rate : {100 - float(self.edt_outlier_rate.text())} %  '

        ws.cell(row=1, column=1).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')    # 색상 노랑
        ws.cell(row=2, column=1).fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')    # 빨강

        ws.cell(row=1, column=2).font = Font(bold=True)     # 글씨 굵게
        ws.cell(row=2, column=2).font = Font(bold=True)
        ws.cell(row=1, column=7).font = Font(bold=True)
        ws.cell(row=2, column=7).font = Font(bold=True)

        wb.save(filename=f'{loc}/{xlsx}')


class Vol_Template:
    def __init__(self):
        super().__init__()
        # self.error_number_lbl = list    # label id 폴더 안의 파일 개수가 맞지 않는 id list
        self.pre_diff = list    # perdict, label 차집합 lsit
        self.lbl_diff = list
        self.accuracy_aver_std = pd.DataFrame()
        self.sheet_aver_std = pd.DataFrame()
        self.error_rate_aver_std = pd.DataFrame()

        self.thin_border = Border(left=borders.Side(style='thin'),
                                  right=borders.Side(style='thin'),
                                  top=borders.Side(style='thin'),
                                  bottom=borders.Side(style='thin'))
        self.blue_color = PatternFill(start_color='b3d9ff', end_color='b3d9ff', fill_type='solid')
        self.green_color = PatternFill(start_color='c1f0c1', end_color='c1f0c1', fill_type='solid')
        self.red_color = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
        self.yellow_color = PatternFill(start_color='ffffb3', end_color='ffffb3', fill_type='solid')
        self.yellow_color2 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.gray_color2 = PatternFill(start_color='e0e0eb', end_color='e0e0eb', fill_type='solid')
        self.gray_color = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
        self.blue_color2 = PatternFill(start_color='ccf5ff', end_color='ccf5ff', fill_type='solid')
        self.outlier_color = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        self.white_color = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    def label(self, loc: str):  # 정답 기준
        # 폴더에 각각 존재하지 않고 하나의 폴더에 전부 있을 때
        logger.info('Label Data Transform Start')
        label_dict = {}
        label_id_list = os.listdir(loc)  # 환자 list

        for i in label_id_list:
            if '.' in i:  # 폴더가 아닌 확장자인 경우 제외
                logger.error(f'폴더가 아닌 파일이 존재합니다. File Name : {i}')
                continue
            data_list = os.listdir(fr'{loc}\{i}')  # air, hts, sts 순서

            data = [-999999, -999999, -999999]  # -999999 로 표현해서 label 폴더에 air,sts,hts 중 없는게 있을 경우 이상치로 만들어 삭제해서 없는 이미지 판단함
            if len(data_list) > 3 and '.png' not in data_list:  # png 파일이 아니거나 3개 초과일 때 제외, 3개 미만인 경우에는 진행하고 결측치로 표현한다.
                logger.error(f'파일 구성이 올바르지 않습니다. File Name : {i} - {data_list}')
                continue
            for j in range(len(data_list)):

                if 'hts' in data_list[j]:  # hts
                    data[2] = (int(re.sub(r'[^0-9]', '', data_list[j])) - 128)/2 + 192  # 정규 표현식으로 문자열 제거, 정답을 ai 기준으로 맞추기 위해서 128씩 더하고 뺌
                elif 'sts' in data_list[j]:  # sts
                    data[1] = int(re.sub(r'[^0-9]', '', data_list[j])) - 192
                elif 'air' in data_list[j]:
                    data[0] = int(re.sub(r'[^0-9]', '', data_list[j])) / 2

            label_dict[i] = data  # dict 에 추가

        df_label = pd.DataFrame(label_dict, index=['Air', 'Soft Tissue', 'Hard Tissue'])
        print(df_label)
        logger.info('label data transform end')
        return df_label

    # 폴더안에 들어있을때
    # def predict(self, loc: str):  # ai 데이터 기준
    #     logger.info('predict data transform start')
    #
    #     predict_dict = {}
    #     predict_id_list = os.listdir(loc)  # 환자 list
    #
    #     for i in predict_id_list:
    #         if '.' in i:  # 폴더가 아닌 확장자가 들어간 파일인 경우 제외
    #             logger.error(f'폴더가 아닌 파일이 존재합니다. File Name : {i}')
    #             continue
    #
    #         data_list = os.listdir(f'{loc}/{i}')  # txt 하나, 2개 이상 일 때 잘못된 것으로 나중에 error 코드 추가
    #
    #         data = [0, 0, 0]
    #         if len(data_list) > 1 and '.dat' not in data_list:  # .dat 파일이 아닌 경우, 폴더안에 파일이 2개 이상인 경우 제외
    #             logger.error(f'파일 구성이 올바르지 않습니다. File Name : {i} - {data_list}')
    #             continue
    #
    #         for j in range(len(data_list)):
    #             txt = open(f'{loc}/{i}/{data_list[j]}', 'r')
    #             lines = txt.readlines()  # txt 한줄씩 읽기
    #
    #             for k in range(len(lines)):
    #                 line = lines[k].split(',')  # air,0.35156316 형식, 숫자만 남김
    #                 line_float = line[1].split('\n')
    #
    #                 if float(line_float[0]) >= 1:
    #                     logger.error(f'predict {j} 폴더의 data 값이 올바르지 않습니다 {line}')  # dat 1이상 값은 없음 error 추가
    #                     data[k] = -999999
    #
    #                 else:
    #                     data[k] = float(line_float[0])
    #
    #         predict_dict[i] = data  # dict 에 추가
    #
    #     df_predict = pd.DataFrame(predict_dict, index=['Air', 'Hard Tissue', 'Soft Tissue'])
    #     logger.info('Label Data Transform End')
    #
    #     return df_predict

    # data 가 전체 밖에 있을 때
    def predict(self, loc: str):  # ai 데이터 기준
        logger.info('predict data transform start')

        predict_dict = {}
        predict_id_list = os.listdir(loc)  # 환자 list

        for j in range(len(predict_id_list)):
            data = [0, 0, 0]
            txt = open(f'{loc}/{predict_id_list[j]}', 'r')
            lines = txt.readlines()  # txt 한줄씩 읽기

            for k in range(len(lines)):
                line = lines[k].split(',')  # air,0.35156316 형식, 숫자만 남김
                line_float = line[1].split('\n')

                if float(line_float[0]) >= 1:
                    logger.error(f'predict {j} 폴더의 data 값이 올바르지 않습니다 {line}')  # dat 1이상 값은 없음 error 추가
                    data[k] = -999999

                else:
                    data[k] = float(line_float[0])

            predict_dict[predict_id_list[j].split('.')[0]] = data  # dict 에 추가
        df_predict = pd.DataFrame(predict_dict, index=['Air', 'Soft Tissue', 'Hard Tissue'])
        df_predict = df_predict*256
        print(df_predict)
        logger.info('Label Data Transform End')
        return df_predict

    def percent(self, lbl: pd.DataFrame, pre: pd.DataFrame, *args) -> pd.DataFrame:  # 퍼센티지의 오차, 셩공률 에 대한 결과 값 함수
        logger.info(f'{args[0]} - Data Transform Start')

        # lbl = 장수 -> 소수점 , pre = 소수점 -> 소수점
        if 'accuracy' in args:
            result_percent = 100-abs(lbl - pre)/256*100
            print(result_percent)
            # lbl_percent = lbl.div(256)  # 나누기
            # lbl_percent = lbl_percent.mul(100)  # 곱
            # pre = pre.mul(100)
            # lbl_percent = lbl_percent.sub(100)
            # pre = pre.sub(100)
            # result_percent = abs(abs(lbl_percent - pre).sub(100))  # 성공 퍼센티지에 대한 결과 lbl-pre 음수 되는 경우 절대값, 100 뺄셈 절대값
        else:
            # lbl_percent = lbl.div(256)
            # lbl_percent = lbl_percent.mul(100)
            # pre = pre.mul(100)
            # result_percent = abs(lbl_percent - pre)  # 오차 퍼센티지에 대한 결과
            result_percent = abs(lbl - pre)/256*100

        result_percent = result_percent[result_percent < 10000]
        result_percent = result_percent.dropna(how='all', axis='columns')

        logger.info(f'{args[0]} - Data Transform End')
        return result_percent

    def sheet(self, lbl: pd.DataFrame, pre: pd.DataFrame) -> pd.DataFrame:  # 장수 차이에 대한 결과 값
        result_sheet = abs(lbl - pre)  # 장수에 대한 결과
        result_sheet = result_sheet[result_sheet < 10000]

        result_sheet = result_sheet.dropna(how='all', axis='columns')

        return result_sheet

    # 결과 값을 토대로 std, aver 를 outlier 적용 한 값도 같이 2개 생성
    def percent_sheet_result(self, outlier, method: pd.DataFrame, *args: str):  # method : 함수 return 값
        logger.info(f'"{args[0]}" - make average, std with outlier dataframe start')

        result = method
        if 'accuracy' in args:
            result_outlier = result[result > outlier]  # outlier 값을 버림 ( nan 으로 만듬 ) // outlier 이상의 값만 남김
        else:
            result_outlier = result[result < outlier]  # outlier 값을 버림 ( nan 으로 만듬 )
        result_average = result.mean(axis=1)  # 평균
        result_std = result.std(axis=1, ddof=0)  # 표준편차

        result_outlier_average = result_outlier.mean(axis=1)
        result_outlier_std = result_outlier.std(axis=1, ddof=0)

        result_aver_std = pd.DataFrame()  # 환자 측정과 평균,표준편차 df 나눔
        result = result.transpose()  # column, row 전환
        result_aver_std.insert(0, 'Remove_Outlier_Std', result_outlier_std)
        result_aver_std.insert(0, 'Remove_Outlier_Aver', result_outlier_average)
        result_aver_std.insert(0, 'Std', result_std)
        result_aver_std.insert(0, 'Aver', result_average)

        result = round(result, 6)  # 소수점 자리수 6
        result_aver_std = round(result_aver_std, 6)  # 소수점 자리수 6
        result_aver_std = result_aver_std.fillna(0)
        logger.info(f'"{args[0]}" - make average, std with outlier dataframe end')

        return result, result_aver_std

    # 각각의 폴더에만 존재할 때, label 폴더의 개수가 맞지 않을 때
    # def pre_lbl_compare(self, label_loc: str, predict_loc: str):
    #
    #     label_id_list = os.listdir(label_loc)  # 환자 list
    #     predict_id_list = os.listdir(predict_loc)  # 환자 list
    #     self.lbl_diff = list(set(label_id_list) - set(predict_id_list))  # 차집합
    #     self.pre_diff = list(set(predict_id_list) - set(label_id_list))
    #     self.error_number_lbl = []
    #
    #     for i in label_id_list:
    #         if '.' in i:    # 폴더가 아닌 확장자인 경우에 제외
    #             pass
    #
    #         else:
    #             number_lbl = os.listdir(f'{label_loc}/{i}')
    #             if len(number_lbl) != 3:
    #                 self.error_number_lbl.append(i)
    #
    #     if len(self.lbl_diff) != 0 or len(self.pre_diff) != 0 or len(self.error_number_lbl) != 0:
    #         logger.error(f'label 폴더에 {self.pre_diff}가 없습니다. predict 폴더에 {self.lbl_diff}가 없습니다. ')
    #         logger.error(f'label 의 개수가 맞지 않는 폴더가 존재합니다. {self.error_number_lbl} ')

    # loc : 엑셀 위치, xlsx : 엑셀명, lbl : label dataframe, pre : predict dataframe, percent_outlier : 퍼센티지 이상치, sheet_outlier :  장수 이상치
    def to_xlsx(self, loc: str, xlsx: str, lbl: pd.DataFrame, pre: pd.DataFrame, error_outlier: float, sheet_outlier: int):  # 엑셀 생성, 결과값 삽입
        logger.info('make xlsx start')

        writer = pd.ExcelWriter(f'{loc}/{xlsx}', engine='openpyxl')  # pandas 엑셀 작성
        # 결과 값 불러옴(pd.dataframe 구성)
        accuracy_result, self.accuracy_aver_std = self.percent_sheet_result(100 - error_outlier, self.percent(lbl, pre, 'accuracy'), 'accuracy')  # 성공률
        error_rate_result, self.error_rate_aver_std = self.percent_sheet_result(error_outlier, self.percent(lbl, pre, 'error rate'), 'error rate')  # 오차율
        sheet_result, self.sheet_aver_std = self.percent_sheet_result(sheet_outlier, self.sheet(lbl, pre), 'sheet')  # 장수 차이

        # df 엑셀에 입력
        accuracy_result.to_excel(writer, startcol=0, startrow=3, sheet_name='보고용')
        self.accuracy_aver_std.to_excel(writer, startcol=5, startrow=3, sheet_name='보고용')
        error_rate_result.to_excel(writer, startcol=0, startrow=3, sheet_name='분석용')
        self.error_rate_aver_std.to_excel(writer, startcol=5, startrow=3, sheet_name='분석용')
        sheet_result.to_excel(writer, startcol=12, startrow=3, sheet_name='분석용')
        self.sheet_aver_std.to_excel(writer, startcol=17, startrow=3, sheet_name='분석용')
        writer.close()

        logger.info('make xlsx end')

    # 이미지 폴더 생성
    def create_img_folder(self, loc: str, file_name: str) -> str:  # 폴더 생성, 위치 값 출력
        img_folder = f'{loc}/{file_name}_graph_image'
        os.mkdir(img_folder)  # 폴더 생성

        return img_folder

    # sheet1 오차율, 장수 차이에 색상, none 값 적용
    def accept_outlier_error_sh1(self, ws, row, column, error, outlier):
        if ws.cell(row=row, column=column).value is None:
            ws.cell(row=row, column=column).value = 'None'
            ws.cell(row=row, column=column).fill = self.gray_color2

        elif ws.cell(row=row, column=column).value >= error:  # error 가 outlier 보다 범위가 크기 떄문에 상위에 있음
            ws.cell(row=row, column=column).fill = self.yellow_color2

            if ws.cell(row=row, column=column).value >= outlier:
                ws.cell(row=row, column=column).fill = self.outlier_color

    # sheet2 성공률에 색상, none 값 적용
    def accept_outlier_error_sh2(self, ws, row, column, error, outlier):
        if ws.cell(row=row, column=column).value is None:
            ws.cell(row=row, column=column).value = 'None'
            ws.cell(row=row, column=column).fill = self.gray_color2

        elif ws.cell(row=row, column=column).value <= error:
            ws.cell(row=row, column=column).fill = self.yellow_color2

            if ws.cell(row=row, column=column).value <= outlier:
                ws.cell(row=row, column=column).fill = self.outlier_color

    # sheet1 에 스타일 적용
    def sheet1_xlsx_style(self, loc: str, xlsx: str, error_outlier: float, sheet_outlier: int, error_rate: float, sheet_error: int, image_folder_loc: str,
                          file_name: str):
        logger.info('Sheet1 Apply Style Start')

        wb = openpyxl.load_workbook(filename=f'{loc}/{xlsx}')
        ws = wb['분석용']

        # if len(self.lbl_diff) != 0 or len(self.pre_diff) != 0 or len(self.error_number_lbl) != 0:    # sheet 1 에는 측정되지 않는 값들 입력함
        #     ws.cell(row=1, column=16).value = f'label 폴더에 {self.pre_diff}가 없습니다. predict 폴더에 {self.lbl_diff}가 없습니다. '
        #     ws.cell(row=1, column=16).font = Font(bold=True)
        #     ws.cell(row=2, column=16).value = f'label 의 개수가 맞지 않는 폴더가 존재합니다. {self.error_number_lbl} '
        #     ws.cell(row=2, column=16).font = Font(bold=True)

        ws.cell(row=4, column=1).value = 'ID/오차율'
        ws.cell(row=4, column=6).value = 'Error Rate'
        ws.cell(row=4, column=6).font = Font(bold=True)
        ws.cell(row=4, column=13).value = 'ID/오차장수'
        ws.cell(row=4, column=18).value = 'Sheet'
        ws.cell(row=4, column=18).font = Font(bold=True)

        # patient 결과 outlier, error 색상 적용 시킴
        for row in range(5, ws.max_row + 1):

            for column in range(2, 5):  # error 범위
                self.accept_outlier_error_sh1(ws, row, column, error_rate, error_outlier)

            for column in range(14, 17):  # sheet 범위
                self.accept_outlier_error_sh1(ws, row, column, sheet_error, sheet_outlier)

            for column in [7, 9]:  # error aver_std 범위
                if ws.cell(row=row, column=column).value is None:
                    pass
                else:
                    self.accept_outlier_error_sh1(ws, row, column, error_rate, error_outlier)

            for column in [19, 21]:  # error aver_std 범위
                if ws.cell(row=row, column=column).value is None:
                    pass
                else:
                    self.accept_outlier_error_sh1(ws, row, column, sheet_error, sheet_outlier)

        # air, hard tissue, soft tissue 색상
        for column in range(1, ws.max_column + 1):
            if ws.cell(row=4, column=column).value is not None and ws.cell(row=4, column=column).value != 'ID/오차장수' and ws.cell(row=4, column=column).value != 'ID/오차율':
                ws.cell(row=4, column=column).fill = self.blue_color

        # # patient_id title 색상 뺌
        # ws.cell(row=4, column=1).fill = self.blue_color2
        # ws.cell(row=4, column=1).border = self.thin_border
        # ws.cell(row=4, column=13).fill = self.blue_color2
        # ws.cell(row=4, column=13).border = self.thin_border

        # 테두리 적용 -> 값이 있는 곳에만 적용
        for row in range(4, ws.max_row + 1):
            for column in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=column).value is not None:
                    ws.cell(row=row, column=column).border = self.thin_border

        # patient_id 색상
        for row in range(5, ws.max_row + 1):
            ws.cell(row=row, column=1).fill = self.yellow_color
            ws.cell(row=row, column=13).fill = self.yellow_color

        # air hts sts 색상 적용
        col_number = [6, 18]
        for i in col_number:
            ws.cell(row=5, column=i).fill = self.green_color
            ws.cell(row=6, column=i).fill = self.green_color
            ws.cell(row=7, column=i).fill = self.green_color

        # 이미지 삽입
        self.graph(self.error_rate_aver_std, 'error_rate', '', loc, error_rate, file_name)
        self.graph(self.error_rate_aver_std, 'Remove_Outlier_error_rate', 'Remove_Outlier_', loc, error_rate, file_name)
        self.graph(self.sheet_aver_std, 'sheet', '', loc, sheet_error, file_name)
        self.graph(self.sheet_aver_std, 'Remove_Outlier_sheet', 'Remove_Outlier_', loc, sheet_error, file_name)
        img_list = os.listdir(f'{loc}/{file_name}_graph_image')

        for i in img_list:
            if '.png' in i:
                img = Image(image_folder_loc + f'/{i}')
                if i == 'Remove_Outlier_error_rate.png':
                    ws.add_image(img, 'F23')
                elif i == 'Remove_Outlier_sheet.png':
                    ws.add_image(img, 'R23')
                elif i == 'sheet.png':
                    ws.add_image(img, 'R9')
                elif i == 'error_rate.png':
                    ws.add_image(img, 'F9')

        # 엑셀 필터 적용 ------------------------------------------------------- 문서에는 적용이 안된다고 하는데 동작이 되서 개수가 많아질 경우 확인해야함.
        # https://openpyxl.readthedocs.io/en/stable/filters.html 공식 사이트
        # 필터 자체를 생성 하는 것은 적용되지만, 생성과 동시에 필터를 적용되지 않는다.
        ws.auto_filter.ref = f'A4:D{ws.max_row}'

        # column 사이즈
        # 오차 평균, 표준편차
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14

        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['I'].width = 19
        ws.column_dimensions['J'].width = 19
        # 장수 평균, 표준편차
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 14
        ws.column_dimensions['O'].width = 14
        ws.column_dimensions['P'].width = 14

        ws.column_dimensions['R'].width = 14
        ws.column_dimensions['U'].width = 19
        ws.column_dimensions['V'].width = 19

        for row in ws[3:ws.max_row]:
            for cell in row:
                cell.alignment = Alignment(horizontal='center',vertical='center')

        ws.title = '분석용'
        wb.save(filename=f'{loc}/{xlsx}')

        logger.info('Sheet1 Apply Style End')

    # sheet2 에 스타일 적용
    def sheet2_xlsx_style(self, loc: str, xlsx: str, error_outlier: float, error_rate: float, image_folder_loc: str, file_name: str):
        logger.info('Sheet2 Apply Style Start')

        error_outlier = abs(error_outlier - 100)  # sheet2 는 성공률
        error_rate = abs(error_rate - 100)

        wb = openpyxl.load_workbook(filename=f'{loc}/{xlsx}')
        ws = wb['보고용']

        ws.cell(row=4, column=1).value = 'ID/정확도'
        ws.cell(row=4, column=6).value = 'Accuracy'
        ws.cell(row=4, column=6).font = Font(bold=True)


        # patient 결과 outlier, error 적용 색상
        for row in range(5, ws.max_row + 1):

            for column in range(2, 5):  # error
                self.accept_outlier_error_sh2(ws, row, column, error_rate, error_outlier)

            for column in [7, 9]:  # error aver_std
                if ws.cell(row=row, column=column).value is None:    # 빈 값은 pass
                    pass
                else:
                    self.accept_outlier_error_sh2(ws, row, column, error_rate, error_outlier)

        # air, hard tissue, soft tissue 색상
        for column in range(1, ws.max_column + 1):
            if ws.cell(row=4, column=column).value is not None and ws.cell(row=4, column=column).value != 'ID/정확도':  # 4행 None, patient id 제외
                ws.cell(row=4, column=column).fill = self.blue_color

        # 존재하지 않는 값 제외 하고 테두리 적용
        for row in range(1, ws.max_row + 1):
            for column in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=column).value is not None:
                    ws.cell(row=row, column=column).border = self.thin_border

        # patient_id 색상
        for row in range(5, ws.max_row + 1):
            ws.cell(row=row, column=1).fill = self.yellow_color

        # 색상 적용
        col_number = [6]
        for i in col_number:
            ws.cell(row=5, column=i).fill = self.green_color
            ws.cell(row=6, column=i).fill = self.green_color
            ws.cell(row=7, column=i).fill = self.green_color

        # 이미지 삽입
        self.graph(self.accuracy_aver_std, 'accuracy', '', loc, error_rate, file_name)
        self.graph(self.accuracy_aver_std, 'Remove_Outlier_accuracy', 'Remove_Outlier_', loc, error_rate, file_name)

        img_list = os.listdir(f'{loc}/{file_name}_graph_image')
        for i in img_list:
            if '.png' in i:    # png 파일 일때만
                img = Image(image_folder_loc + f'/{i}')
                if i == 'remove_Outlier_accuracy.png':
                    ws.add_image(img, 'F23')
                elif i == 'accuracy.png':
                    ws.add_image(img, 'F9')

        ws.auto_filter.ref = f'A4:D{ws.max_row}'  # 엑셀 필터 적용

        # column 사이즈
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['I'].width = 19
        ws.column_dimensions['J'].width = 19
        ws.column_dimensions['U'].width = 19
        ws.column_dimensions['V'].width = 19

        for row in ws[3:ws.max_row]:
            for cell in row:
                cell.alignment = Alignment(horizontal='center',vertical='center')

        ws.title = '보고용'
        wb.save(filename=f'{loc}/{xlsx}')

        logger.info('Sheet2 Apply Style End')

    # 그래프 제작
    def graph(self, df: pd.DataFrame, title: str, outlier: str, loc: str, error_line, file_name: str):
        logger.info(f'"{title}" - Make Graph Start')
        graph = df
        graph_dict = graph.to_dict('list')  # dataframe list 제작

        fig = plt.figure(figsize=(5, 3))  # Figure 생성 사이즈
        ax = fig.add_subplot()  # Axes 추가
        colors = ['#c1f0c1', '#c1f0c1', '#c1f0c1']    # 초록색
        xtick_label_position = list(range(len(list(graph.index))))  # x 축에 글시 넣을 위치

        if 'accuracy' in title:    # 성공률 일때 적용
            for j in range(len(list(graph.index))):
                if float(graph_dict[f'{outlier}Aver'][j]) <= float(error_line):    # error_line 값보다 낮으면 색변환 // 성공률 낮으면~
                    colors[j] = '#FFCCCC'  # error 빨강

        else:
            for j in range(len(list(graph.index))):
                if float(graph_dict[f'{outlier}Aver'][j]) >= float(error_line):    # error_line 값보다 낮으면 색변환 // 오차율 높으면~
                    colors[j] = '#FFCCCC'  # error 빨강

        if 'error_rate' in title:  # sheet 와 error 에 따른 축 범위 변경
            plt.ylim([0, error_rate_range])
            # plt.gca().yaxis.set_major_formatter(mticker.FormatStrFormaater('%i %'))
            if 'accuracy' in title:
                plt.ylim([0, accuracy_range])

        elif 'sheet' in title:
            plt.ylim([0, sheet_range])

        plt.xticks(xtick_label_position, list(graph.index))  # x 축에 삽입
        plt.axhline(y=float(error_line), color='red', linestyle='--')  # error 라인 그리기
        bars = plt.bar(xtick_label_position, graph_dict[f'{outlier}Aver'], color=colors, edgecolor='black')  # 그래프 생성
        plt.title(title, fontsize=10)  # 타이틀 입력
        plt.errorbar(x=list(graph.index), y=graph_dict[f'{outlier}Aver'], yerr=np.array(graph_dict[f'{outlier}Std']) / 2, color='black', ecolor='black', fmt='.',
                     alpha=0.5, elinewidth=2)  # 에러바 삽입

        for i, b in enumerate(bars):  # 바에 결과 값 추가
            ax.text(b.get_x() + b.get_width() / 2, b.get_height() / 2, graph_dict[f'{outlier}Aver'][i], ha='center', fontsize=10)

        plt.xticks(rotation=0)
        plt.savefig(f'{loc}/{file_name}_graph_image/{title}.png')
        logger.info(f'"{title}" - Make Graph End')
# if __name__ == "__main__":
#     vol = Vol_Template()
#     xlsx_name = 'result.xlsx'
#     df_lbl = vol.label(r'D:\temp_data\label')
#     df_pre = vol.predict(r'D:\temp_data\predict')
#
#     image_folder = vol.create_img_folder(r'D:\temp_data', 'ff')
#     vol.to_xlsx(fr'D:\temp_data', xlsx_name, df_lbl, df_pre, 5, 15)
#     vol.sheet1_xlsx_style(fr'D:\temp_data', xlsx_name, 5, 13, 3, 7, image_folder, 'ff')
#     vol.sheet2_xlsx_style(fr'D:\temp_data', xlsx_name, 5, 3, image_folder, 'ff')
# fr'D:\temp_data\graph_image'
